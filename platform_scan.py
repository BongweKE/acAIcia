"""
GENE-LINK Platform Scan — Merged & Consolidated Workbook
=========================================================
Merges our Platform Scan v01 with the Platform Ecosystem Scan Tool
contributed by Dominique.

Merge principles
----------------
1.  All information from both sources is preserved.
2.  Rows that cover the same source are merged into one rich row
    (ABSCH is the only direct duplicate across both files).
3.  Content that lives in both regulatory-data tabs and content
    tabs (e.g. Nagoya Protocol text) appears once in the data-
    source sheet and once in II-E with an explicit dedup note.
4.  Dominique's Architecture Notes becomes I-B (UI/UX companion).
5.  Dominique's In-Country Support template is kept as II-F.
6.  Knowledge Assets + E-Learning merge into II-E with section
    dividers; topics already covered in II-A are flagged not
    duplicated.
7.  AI/ML functionalities get a dedicated section in Part III.

Final tab structure
-------------------
  0.  Instructions
  I-A.  Existing Platforms          (38 platforms, both sources)
  I-B.  Platform Architecture Notes (Dominique — UX/tech deep-dive)
  II-A. Regulatory Data Sources
  II-B. Company & Innovation Sources
  II-C. Resource & Biodiversity DBs (adds GenBank, ENA, DDBJ, WDPA…)
  II-D. Southern Institution DBs
  II-E. Content & Knowledge Assets  (e-learning + legal assets, merged)
  II-F. In-Country Support          (empty template from Dominique)
  III.  Functionalities & Requirements (adds AI/ML section)
"""

from __future__ import annotations
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Design tokens
# ---------------------------------------------------------------------------
DARK_GREEN = "1B5E20"
MID_GREEN = "2E7D32"
LIGHT_GREEN = "A5D6A7"
PALE_GREEN = "E8F5E9"
NAVY = "1A237E"
BLUE_MID = "1565C0"
BLUE_PALE = "E3F2FD"
AMBER = "E65100"
AMBER_PALE = "FFF3E0"
PURPLE = "4A148C"
PURPLE_PALE = "F3E5F5"
TEAL = "004D40"
TEAL_PALE = "E0F2F1"
INDIGO = "283593"
INDIGO_PALE = "E8EAF6"
BROWN = "4E342E"
BROWN_PALE = "EFEBE9"
WHITE = "FFFFFF"
LIGHT_GREY = "F5F5F5"
MED_GREY = "EEEEEE"


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)


def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)


def _border(style="thin") -> Border:
    s = Side(style=style, color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)


def _left() -> Alignment:
    return Alignment(wrap_text=True, vertical="top", horizontal="left")


def _center() -> Alignment:
    return Alignment(wrap_text=True, vertical="top", horizontal="center")


def _banner(ws, title: str, subtitle: str) -> None:
    ws.merge_cells("A1:V1")
    c = ws.cell(row=1, column=1, value=title)
    c.font = _font(bold=True, color=WHITE, size=13)
    c.fill = _fill(NAVY)
    c.alignment = _left()
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:V2")
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = _font(italic=True, color=NAVY, size=9)
    c2.fill = _fill(BLUE_PALE)
    c2.alignment = _left()
    ws.row_dimensions[2].height = 16


def _header(ws, headers: list[str], row: int, fill_hex=NAVY) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(fill_hex)
        c.alignment = _center()
        c.border = _border()
    ws.row_dimensions[row].height = 42


def _row(ws, vals: list[Any], r: int, shade=False) -> None:
    bg = LIGHT_GREY if shade else WHITE
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=col, value=v)
        c.font = _font(size=9)
        c.fill = _fill(bg)
        c.alignment = _left()
        c.border = _border()


def _section(ws, label: str, r: int, fill_hex=BLUE_PALE, ncols=22) -> None:
    ws.merge_cells(f"A{r}:{get_column_letter(ncols)}{r}")
    c = ws.cell(row=r, column=1, value=label)
    c.font = _font(bold=True, color=NAVY, size=9)
    c.fill = _fill(fill_hex)
    c.alignment = _left()
    c.border = _border()
    ws.row_dimensions[r].height = 16


def _widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet 0 — Instructions
# ---------------------------------------------------------------------------

INSTRUCTIONS = [
    ("Scan", "Platform Scan — one of four parallel GENE-LINK scans (Partner | "
     "Company | Governance | Platform). This workbook covers the Platform Scan "
     "exclusively and merges contributions from both the GENE-LINK platform team "
     "and Platform Ecosystem Scan Tool."),
    ("Objective", "Survey existing platforms and data sources to design the "
     "GENE-LINK platform's UI/UX, backend data architecture, and operational "
     "rules. Outputs feed directly into the System Requirements Specification (SRS)."),
    ("Part I — Existing Platforms\n(Sheets I-A and I-B)",
     "I-A: 38 platforms covering matchmaking, compliance, capacity development, "
     "marketplace/trust models, collaboration tools, analytics, and UX reference. "
     "I-B: Architecture Notes — UX/technical deep-dives on each platform (onboarding "
     "flows, matching algorithms, design failures to avoid). Both sheets steer "
     "UI/UX design. Cross-reference by Platform Name."),
    ("Part II — Existing Data Sources\n(Sheets II-A through II-F)",
     "II-A: Regulatory ABS data sources. II-B: EU company and innovation databases. "
     "II-C: Resource, biodiversity, and genomic databases (incl. DSI databases: "
     "GenBank, ENA, DDBJ). II-D: Southern institution databases. II-E: Content and "
     "knowledge assets (e-learning courses + legal templates and frameworks — "
     "merged, deduped). II-F: In-country support ecosystem (empty template). "
     "Goal: design the platform backend."),
    ("Part III — Functionalities\n(Sheet III)",
     "Three sections: (1) Access & Discovery, (2) Verification & Due Diligence, "
     "(3) Security, and (4) AI / ML Features. Each row maps to a [SRS-REQ] "
     "statement. Goal: design the platform's operational architecture."),
    ("Relationship to Other Scans",
     "Partner Scan → identifies Southern provider institutions (feeds II-D). "
     "Company Scan → identifies EU user companies (feeds II-B). "
     "Governance Scan → assesses ABS regulatory conditions (feeds II-A). "
     "Where this workbook references those topics, 'Cross-Scan Linkage' columns "
     "flag it. This workbook captures sources from a platform-integration angle."),
    ("Merge Notes",
     "Rows contributed by Dominique carry 'Dominique' in the Added By column. "
     "Rows contributed by the GENE-LINK platform team carry 'Platform Team'. "
     "Merged rows (same source in both contributions) carry 'Both / Merged'. "
     "The ABS Clearing-House is the only direct platform duplicate; it is merged "
     "into a single row in I-A."),
    ("Deduplication Policy",
     "Sources that appear in both a data-source sheet (II-A–D) and the content "
     "tab (II-E) are not removed — they represent different dimensions of the same "
     "source (data API vs. training content). A 'Dedup Note' column in II-E "
     "flags where the data-source dimension is captured elsewhere."),
    ("AI / ML",
     "AI features are captured in a dedicated section at the bottom of Sheet III. "
     "They are not spread across other sheets. Where a data source or platform has "
     "direct AI integration potential, this is noted in the SRS Requirement column "
     "of that row."),
    ("SRS Column Convention",
     "Any row generating a platform design implication has a value in the "
     "'SRS Requirement' column formatted as: [SRS-REQ] <statement>. These are "
     "harvested into the Requirements Specification document."),
    ("Priority Convention",
     "High = required for MVP / pilot. Medium = full launch. Low = future phase."),
    ("Version", "v0.2 — Merged edition (March 2026). Both source workbooks "
     "fully incorporated. Columns marked * require validation via platform "
     "testing and key informant interviews."),
]


# ---------------------------------------------------------------------------
# Sheet I-A — Existing Platforms (38 rows, unified 21-column schema)
# ---------------------------------------------------------------------------
# Columns:
#  0  Platform Name           7  Key Functionalities         14 Commercial Design Features
#  1  URL                     8  Data / Profile Structure     15 Integration / Linkage Potential
#  2  Platform Type/Category  9  Business Model / Funding     16 ABS / GR Specificity
#  3  Status                 10  Key Features for GENE-LINK   17 Priority for Follow-up
#  4  Description            11  Strengths to Learn From      18 Added By
#  5  Geographic Focus       12  Known Limitations            19 SRS Implication
#  6  Primary User Base      13  UI/UX Notes for GENE-LINK

PLAT_COLS = [
    "Platform Name", "URL", "Platform Type / Category", "Status",
    "Description", "Geographic Focus", "Primary User Base",
    "Key Functionalities", "Data / Profile Structure",
    "Business Model / Funding",
    "Key Features Relevant to GENE-LINK",
    "Strengths to Learn From",
    "Known Limitations / Design Failures to Avoid",
    "UI/UX Notes for GENE-LINK",
    "Positioning Opportunity / Gaps GENE-LINK Should Fill",
    "Commercial Design Features to Adopt",
    "Integration / Linkage Potential",
    "ABS / GR Specificity\n(High/Med/Low/None)",
    "Priority for Follow-up",
    "Added By",
    "SRS Implication",
]

# Each row: 21 values matching PLAT_COLS order
PLAT_ROWS = {
    "MATCHMAKING": [
        ["b2match", "https://www.b2match.com",
         "Matchmaking (EU B2B)", "Active",
         "EU-aligned B2B matchmaking enabling structured collaboration between "
         "SMEs, research institutions, and public sector actors.",
         "EU / Global", "SMEs; research institutions; public sector",
         "Profiles; structured matchmaking; meeting scheduling",
         "Structured (offers; requests; expertise tags)",
         "Private SaaS / Subscription",
         "Structured collaboration fields; sector and role filtering; meeting "
         "scheduling layer most relevant to GENE-LINK.",
         "Trusted EU standard for structured collaboration; clear offer/request "
         "profile distinction.",
         "Event-centric lifecycle; dated UX; limited continuous engagement.",
         "Adopt structured offer/request/expertise field model. Extend beyond "
         "event context into always-on ecosystem.",
         "Extend beyond events into continuous ecosystem matchmaking with "
         "ABS-specific filters.",
         "Structured matchmaking fields; collaboration type tagging",
         "High", "None", "High", "Dominique",
         "[SRS-REQ] GENE-LINK matchmaking profiles must implement structured "
         "offer/request/expertise fields and sector/resource-type filtering "
         "modelled on b2match's collaboration schema."],

        ["Grip", "https://www.grip.events",
         "Matchmaking (AI-driven)", "Active",
         "AI-driven matchmaking platform for large-scale B2B networking; uses "
         "behavioural signals alongside profile data.",
         "Global", "Corporates; startups; investors",
         "AI matching; meeting scheduling; recommendation engine",
         "Profiles + behavioural / engagement signals",
         "Private SaaS / Subscription",
         "Mutual-interest matching logic; behavioural signal weighting; "
         "recommendation engine architecture.",
         "Best-in-class UX; mutual interest matching; behavioural signals "
         "improve relevance over time.",
         "Event-focused engagement lifecycle; limited persistence beyond events.",
         "Mutual interest logic (both parties must signal interest before "
         "match is made) is essential for GENE-LINK trust model.",
         "Improve GENE-LINK matchmaking UX beyond basic search.",
         "Recommendation engine; mutual interest signal before connection",
         "Medium–High", "None", "High", "Dominique",
         "[SRS-REQ] GENE-LINK matchmaking engine must implement mutual-interest "
         "logic: a match is only surfaced when both provider and user have "
         "signalled interest, preventing unsolicited outreach."],

        ["Swapcard", "https://www.swapcard.com",
         "Event + Matchmaking (Hybrid)", "Active",
         "Platform combining networking, content, and AI matchmaking; rich "
         "engagement data from content consumption informs recommendations.",
         "Global", "Event ecosystems; corporates",
         "Networking; content; AI matching; rich profile engagement",
         "Rich profiles + session/content engagement data",
         "Private SaaS / Subscription",
         "Content-integrated matchmaking: engagement with content (e.g. "
         "reading an ABS guide) can signal intent for the matching engine.",
         "Strong UX integration of content and networking; engagement-driven "
         "recommendation improvement.",
         "Episodic engagement bound to event lifecycle.",
         "Consuming ABS learning content can be a matching signal — users "
         "who read a Brazil ABS guide are likely sourcing from Brazil.",
         "Continuous engagement beyond events; use content consumption as "
         "a matching signal.",
         "Content + networking integration; behavioural matching",
         "Medium", "None", "Medium", "Dominique",
         "[SRS-REQ] Platform should track content consumption (e.g. reading a "
         "country ABS guide) as an implicit intent signal that improves "
         "matchmaking relevance."],

        ["WIPO GREEN", "https://www.wipo.int/wipogreen",
         "Matchmaking (Technology Transfer)", "Active",
         "WIPO green technology matchmaking platform connecting technology "
         "owners with technology seekers. Closest multilateral analogue "
         "to GENE-LINK in mandate.",
         "Global", "Companies; researchers; technology transfer offices",
         "Technology listings; structured matchmaking; needs posting",
         "Technology and need profiles",
         "Multilateral / Donor-funded",
         "Most directly analogous multilateral matchmaking platform to "
         "GENE-LINK. Its low engagement rate is the key cautionary lesson.",
         "Relevant thematic model; multilateral legitimacy; structured "
         "technology-offer/need profiles.",
         "Low engagement; minimal follow-through to actual deals; weak "
         "transaction and compliance layer.",
         "GENE-LINK must solve the engagement problem WIPO GREEN has not: "
         "passive listing is insufficient; active curation and guided "
         "pathways are essential.",
         "Add active deal facilitation, ABS compliance layer, and "
         "guided user journey on top of listing model.",
         "Structured technology/need listings; multilateral positioning",
         "High", "Medium", "Very High", "Dominique",
         "[SRS-REQ] GENE-LINK must avoid WIPO GREEN's passive-listing failure: "
         "platform must include active guided pathways, notifications, and "
         "deal-facilitation features beyond static profile listing."],

        ["ABS Connect", "https://absconnect.org",
         "Matchmaking (ABS-specific)", "Active",
         "Online brokerage connecting providers and users of genetic resources "
         "with ABS-compliance awareness.",
         "Global (ABS-specific)", "Researchers; companies; ABS practitioners",
         "Profile matching by resource type; direct messaging; ABS framing",
         "Structured provider/user profiles",
         "Free (limited) / Institutional",
         "Only dedicated ABS-aware matchmaking platform; resource-type and "
         "sector-based filtering; ABS-specific framing.",
         "Simple, ABS-specific interface; resource-type filtering logic "
         "directly applicable to GENE-LINK.",
         "Limited to introductory contact; no compliance workflow, MTA "
         "automation, or Southern institution capacity profiling.",
         "GENE-LINK is the evolution of this model: add compliance workflow, "
         "MTA tooling, and verified institutional profiles.",
         "Extend beyond contact introduction into full deal facilitation.",
         "Resource-type filtering; ABS-aware profile structure",
         "High", "High", "High", "Platform Team",
         "[SRS-REQ] GENE-LINK must replicate ABS Connect's resource-type and "
         "sector-based filtering as the minimum matching dimension, then layer "
         "compliance workflows and MTA tooling on top."],

        ["WIPO MATCH", "https://www.wipo.int/en/web/wipo-match",
         "Matchmaking (IP / Technology Transfer)", "Active",
         "IP-based technology transfer matchmaking connecting technology owners "
         "with technology seekers; licensing and collaboration brokerage.",
         "Global", "Technology owners; licensees; universities; companies",
         "Structured offer/seek profiles; sector tags; geographic filters; "
         "direct contact facilitation",
         "Structured offer/seek profiles",
         "Free (registration required)",
         "Clean two-sided marketplace with explicit offer/seek distinction; "
         "sector tagging; geographic filtering.",
         "Clear two-sided offer/seek structure directly applicable to "
         "GENE-LINK's provider/user model.",
         "Not ABS-aware; no genetic resource or compliance specificity.",
         "GENE-LINK provider/user profile structure should follow the offer/"
         "seek model with resource-specific fields added.",
         "Add ABS compliance, genetic resource specificity, and Southern "
         "institution capacity fields.",
         "Offer/seek two-sided profile model",
         "High", "Low", "Medium", "Platform Team",
         "[SRS-REQ] GENE-LINK profiles must implement a clear two-sided "
         "offer/seek structure: Southern institutions offer resources and "
         "capacity; EU companies seek specific resource types and use cases."],

        ["Nature Research Partnerships",
         "https://partnerships.nature.com",
         "Matchmaking (Academic–Industry)", "Active",
         "Curated academic-industry collaboration brokerage by Springer Nature.",
         "Global (EU/US heavy)",
         "Academic researchers; corporate R&D teams",
         "Curated matchmaking; quality-screened profiles; sector outreach",
         "Curated researcher and company profiles",
         "Paid (corporate)",
         "High-quality editorial curation model; demonstrates that human "
         "curation alongside algorithmic matching is sometimes necessary.",
         "Human curation ensures quality that algorithmic matching alone "
         "cannot achieve for sensitive partnerships.",
         "Not ABS-aware; corporate framing poorly suited to Southern "
         "institution context; no compliance layer.",
         "GENE-LINK may need human-curated shortlisting layer for high-value "
         "partnerships, especially where ABS sensitivity is high.",
         "Add ABS compliance, Southern institution context, and open access "
         "pricing model.",
         "Editorial quality curation model",
         "Low", "None", "Low", "Platform Team", ""],
    ],

    "COMMUNITY & COLLABORATION": [
        ["Discourse", "https://www.discourse.org",
         "Community / Forum", "Active",
         "Open-source platform for structured knowledge exchange, threaded "
         "discussion, and community governance.",
         "Global", "Communities; organisations; NGOs",
         "Threaded discussions; tagging; search; moderation tools",
         "User + contribution profiles; topic-based structure",
         "Open source / Subscription or self-hosted",
         "Persistent, searchable knowledge base; moderation tools; governance "
         "layer for community management.",
         "Persistent knowledge — conversations become searchable assets. "
         "Community governance tools. Strong API for integration.",
         "Low engagement without active facilitation; requires moderation "
         "investment.",
         "GENE-LINK's knowledge-dialogue layer (e.g. ABS Q&A, partner "
         "discussions) should follow Discourse's persistent, searchable "
         "thread model rather than ephemeral chat.",
         "Long-term dialogue layer anchoring institutional knowledge.",
         "Threaded searchable knowledge; tagging; moderation governance",
         "High", "None", "High", "Dominique",
         "[SRS-REQ] Platform knowledge-dialogue layer must implement a "
         "persistent, searchable forum with topic tagging and moderation "
         "controls — not ephemeral chat — so that ABS Q&A and partner "
         "discussions become reusable knowledge assets."],

        ["Microsoft Teams", "https://www.microsoft.com/teams",
         "Secure Collaboration", "Active",
         "Enterprise communication and collaboration platform; primary secure "
         "comms tool for EU institutional actors.",
         "Global", "Corporates; institutions; governments",
         "Messaging; video calls; file sharing; channel governance",
         "Organisation-linked identities; permission-based channels",
         "Microsoft / Subscription",
         "Enterprise-grade security; EU-trusted; strong institutional "
         "adoption among GENE-LINK target EU company users.",
         "Secure; trusted; strongest EU institutional adoption.",
         "Not marketplace-specific; no matchmaking or transaction layer; "
         "fragmented experience across many channels.",
         "GENE-LINK's secure communication layer for active partnerships "
         "should integrate with or replicate Teams' permission model. "
         "Do not build a custom comms tool from scratch.",
         "Secure interaction layer for active deal communications.",
         "Enterprise communication model; channel-based permissions",
         "High", "None", "High", "Dominique",
         "[SRS-REQ] Platform secure communication module for active "
         "partnerships must integrate with MS Teams (and/or Rocket.Chat "
         "as open-source alternative) rather than building a custom "
         "messaging layer."],

        ["Rocket.Chat", "https://www.rocket.chat",
         "Secure Messaging (Open Source)", "Active",
         "Open-source self-hosted messaging for high-security, data-sovereign "
         "environments. Alternative to MS Teams for institutions with strict "
         "data residency requirements.",
         "Global", "Governments; NGOs; institutions",
         "Messaging; channels; self-hosted; API-driven",
         "User accounts + channels; self-hosted data",
         "Self-hosted / Subscription",
         "Data sovereignty for Southern institution partners who cannot "
         "use Microsoft cloud services.",
         "Data sovereignty; self-hosting option; strong API flexibility.",
         "Lower UX polish than Teams; smaller ecosystem.",
         "Offer as an alternative comms layer for Southern institution "
         "partners who require on-premise or regional hosting.",
         "High-security communications for data-sovereign environments.",
         "Self-hosting; data sovereignty; open-source extensibility",
         "Medium", "None", "Medium", "Dominique",
         "[SRS-REQ] Platform must offer a data-sovereign communication "
         "option (e.g. Rocket.Chat) for Southern institution partners "
         "who cannot use US-cloud-based messaging tools."],

        ["GitHub", "https://github.com",
         "Collaborative Development", "Active",
         "Platform for collaborative software development, version control, "
         "and open project coordination.",
         "Global", "Developers; open-source communities",
         "Repositories; version control; issue tracking; collaboration",
         "Code repositories + contributor profiles",
         "Private / Open source / Subscription",
         "Transparent collaboration model; version-controlled asset "
         "management; issue tracking as a workflow tool.",
         "Radical transparency in collaborative development; versioning "
         "creates audit trails; issue tracking models task management.",
         "High technical barrier for non-developers.",
         "GENE-LINK's platform codebase should be on GitHub. Versioned "
         "document management (e.g. MTA templates) could follow a "
         "Git-inspired model.",
         "Model version-controlled collaboration for platform assets.",
         "Version control; transparent collaboration; issue tracking",
         "Medium", "None", "Medium", "Dominique", ""],
    ],

    "DOCUMENT & DEAL INFRASTRUCTURE": [
        ["DocSend", "https://www.docsend.com",
         "Secure Document Sharing", "Active",
         "Secure document sharing with controlled access, view tracking, and "
         "analytics. Used for pre-agreement stage document exchange.",
         "Global", "Startups; corporates",
         "Document sharing; access analytics; link-based control",
         "Document-centric; access tracked per viewer",
         "Private SaaS / Subscription",
         "Controlled document visibility with viewer-level tracking; "
         "access revocation; simple link-based sharing.",
         "Simple yet powerful: granular access control without requiring "
         "the recipient to create an account.",
         "Limited collaborative workflow; no versioning; not for "
         "multi-party agreement execution.",
         "GENE-LINK's pre-deal document sharing (e.g. sharing a draft MTA "
         "for review) should implement DocSend-style access tracking and "
         "revocation.",
         "Bridge between first contact and formal agreement execution.",
         "Access tracking; link-based controlled sharing; revocation",
         "High", "None", "High", "Dominique",
         "[SRS-REQ] Platform document vault must implement viewer-level "
         "access tracking and link revocation for pre-agreement documents, "
         "modelled on DocSend's access analytics."],

        ["Intralinks", "https://www.intralinks.com",
         "Virtual Data Room", "Active",
         "High-security enterprise virtual data room for deal management and "
         "sensitive document sharing with full audit trails.",
         "Global", "Corporates; investors; M&A teams",
         "Data rooms; permission hierarchies; audit trails; NDAs",
         "Document-centric with strict permission tiers",
         "Enterprise SaaS / Subscription",
         "Immutable audit trails; high-security data rooms for complex "
         "multi-party deals; NDA workflow integration.",
         "Top-tier security and audit trail model for high-value deals.",
         "Complex and costly; over-engineered for early-stage or "
         "research-level partnerships.",
         "GENE-LINK should adopt Intralinks' audit trail and permission "
         "hierarchy model for the compliance document vault, but only "
         "for Verified-tier users and high-value commercial deals.",
         "Support high-value commercial ABS deals in Phase 2.",
         "Audit trails; permission hierarchy; NDA workflow",
         "Medium–High", "None", "Medium", "Dominique",
         "[SRS-REQ] Platform compliance document vault must implement "
         "immutable audit trails following Intralinks' model. Phase 1 "
         "can use a simplified version; Phase 2 should support full "
         "virtual data room for commercial ABS deals."],
    ],

    "MARKETPLACE & TRUST MODELS": [
        ["Airbnb", "https://www.airbnb.com",
         "Trust Marketplace", "Active",
         "Platform connecting hosts and guests with strong trust mechanisms: "
         "identity verification, dual reviews, and financial escrow.",
         "Global", "Individuals; SMEs",
         "Listings; dual reviews; identity verification; booking; escrow",
         "Profiles + reviews + identity verification",
         "Private platform / Transaction fees",
         "Dual review system (both parties rate each other); identity "
         "verification before booking; progressive trust accumulation.",
         "Best-in-class trust signal architecture: identity verification "
         "+ dual reviews prevents gaming. Reputation is mutual.",
         "Platform dependency; price commoditisation; reviews can be "
         "gamed at scale.",
         "GENE-LINK's trust layer should implement dual review "
         "(EU company reviews Southern provider AND vice versa) plus "
         "identity verification as a precondition for any financial "
         "transaction.",
         "Model trust architecture for ABS partnership execution.",
         "Dual review system; identity verification; progressive trust",
         "Very High", "None", "Very High", "Dominique",
         "[SRS-REQ] Platform must implement a dual review system where "
         "both EU user and Southern provider rate each other after "
         "partnership completion, creating mutual accountability."],

        ["Fiverr", "https://www.fiverr.com",
         "Marketplace (Reputation-based)", "Active",
         "Service marketplace with embedded reputation, rating, and "
         "retention systems that prevent disintermediation.",
         "Global", "Freelancers; SMEs",
         "Search; ratings; service tiers; transaction within platform",
         "Service profiles + ratings + completion history",
         "Private platform / Transaction fees",
         "Reputation accumulation within the platform; tiered seller "
         "status creates retention incentive.",
         "Prevents disintermediation: reputation is non-portable, "
         "so users have incentive to transact through the platform.",
         "Price competition at low end; commoditises quality; "
         "complex gig-economy dynamics may not suit institutional context.",
         "GENE-LINK's reputation system must create non-portable "
         "value (platform-internal trust score) to prevent partners "
         "from taking deals off-platform.",
         "Build trust-based partnerships that stay within the platform.",
         "Reputation accumulation; tiered trust; anti-disintermediation",
         "Medium", "None", "High", "Dominique",
         "[SRS-REQ] Platform must implement a non-portable, platform-"
         "internal reputation/trust score that creates incentive to "
         "execute all transactions through the platform."],

        ["Upwork", "https://www.upwork.com",
         "Marketplace (Structured Contracting)", "Active",
         "Platform for structured freelance contracts with escrow, "
         "milestone payments, and dispute resolution.",
         "Global", "Professionals; businesses",
         "Bidding; contracts; escrow; milestone payments; dispute resolution",
         "Profiles + work history + ratings",
         "Private platform / Transaction fees",
         "Escrow-based trust for payment; milestone-based contract "
         "structure; dispute resolution as safety net.",
         "Escrow payment model builds financial trust between parties "
         "who have not worked together before.",
         "Complex onboarding; assumes freelance dynamics; "
         "not directly applicable to institutional research partnerships.",
         "GENE-LINK benefit-sharing payment mechanisms could use an "
         "escrow model for initial transactions before parties have "
         "established direct trust.",
         "Model structured agreements with financial escrow for "
         "first-time transactions.",
         "Escrow payments; milestone contracts; dispute resolution",
         "Medium", "None", "High", "Dominique",
         "[SRS-REQ] Platform should support milestone-based benefit-sharing "
         "payment schedules, with escrow functionality planned for Phase 2 "
         "to reduce financial risk for first-time partnerships."],

        ["AngelList", "https://angel.co",
         "Startup Investment Marketplace", "Active",
         "Platform connecting startups with investors and talent; "
         "strong deal visibility and investment matching.",
         "Global", "Startups; investors",
         "Startup profiles; investment deals; talent matching",
         "Startup profiles + funding data",
         "Private platform / Fees",
         "Investment matching logic; deal pipeline visibility; "
         "structured startup profile with traction metrics.",
         "Investment matching and deal pipeline model relevant for "
         "GENE-LINK's higher-value commercial partnership flows.",
         "Niche (startup/investor ecosystem); not directly applicable "
         "to research/ABS context.",
         "GENE-LINK's partnership pipeline view could follow AngelList's "
         "deal visibility model: providers see which EU companies "
         "are interested; companies see the pipeline of available resources.",
         "Model partnership pipeline visibility and deal tracking.",
         "Deal visibility; pipeline tracking; traction metrics",
         "Medium", "None", "High", "Dominique", ""],
    ],

    "PROFESSIONAL & ACADEMIC NETWORKS": [
        ["LinkedIn", "https://www.linkedin.com",
         "Professional Network", "Active",
         "Professional identity and networking platform; de facto global "
         "standard for professional credibility.",
         "Global", "Professionals; companies",
         "Profiles; messaging; feeds; company pages; endorsements",
         "Professional identity profiles (roles; experience; skills; network)",
         "Private platform / Ads + subscriptions",
         "Professional identity benchmark; progressive profile completion; "
         "endorsements as credibility signals.",
         "Identity + credibility signals; progressive profile completion "
         "is the strongest onboarding model.",
         "No transaction layer; no ABS specificity; weak for "
         "non-Western professional contexts.",
         "GENE-LINK profiles should follow LinkedIn's progressive completion "
         "model — users start with minimal data and are incentivised to "
         "enrich their profiles over time.",
         "Benchmark professional profiles and credibility architecture.",
         "Endorsements; progressive profile completion; identity credibility",
         "High", "None", "High", "Dominique",
         "[SRS-REQ] Platform profiles should implement progressive completion "
         "incentives modelled on LinkedIn: users gain access to more features "
         "as their profile completeness score increases."],

        ["ResearchGate", "https://www.researchgate.net",
         "Academic Research Network", "Active",
         "Platform connecting researchers and publications; academic "
         "credibility layer.",
         "Global", "Researchers; academics",
         "Researcher profiles; publications; project pages; Q&A",
         "Research profiles + publication records",
         "Private / Freemium",
         "Academic credibility via publication records; ORCID integration "
         "path; relevant for Southern institution researcher profiles.",
         "Academic credibility anchor for researcher profiles; "
         "bridges research outputs to professional identity.",
         "Limited private sector integration; no ABS specificity; "
         "restricted API access.",
         "Southern institution researcher profiles in GENE-LINK should "
         "link to ResearchGate (or ORCID) for academic credibility "
         "verification without rebuilding the publication layer.",
         "Bridge academic research credibility to commercial partnership.",
         "Publication-linked credibility; research project profiles",
         "Medium", "None", "Medium", "Dominique",
         "[SRS-REQ] Southern institution researcher profiles must support "
         "ORCID integration for identity and publication-record linkage; "
         "ResearchGate links are a secondary credibility signal."],

        ["CORDIS", "https://cordis.europa.eu",
         "EU Research Platform", "Active",
         "EU platform for funded research results; comprehensive database "
         "of Horizon Europe and FP projects.",
         "EU", "Researchers; institutions; companies",
         "Project database; results; funding data; partner search",
         "Structured EU project profiles",
         "EU / Public funding",
         "EU project legitimacy; Horizon Europe partner discovery; "
         "identifies EU research actors active in relevant fields.",
         "Strong EU legitimacy; systematic coverage of EU-funded projects "
         "relevant to bio-innovation.",
         "Static and non-interactive; no transaction or communication layer.",
         "GENE-LINK should integrate with CORDIS to identify EU research "
         "teams and companies active in bio-innovation projects — "
         "a warm prospecting list for EU user onboarding.",
         "Integrate EU research knowledge; identify EU prospecting targets.",
         "Structured EU project data; research consortium profiles",
         "High", "None", "High", "Dominique",
         "[SRS-REQ] Platform EU company onboarding should cross-reference "
         "CORDIS to identify companies and institutions active in relevant "
         "EU-funded research, enabling targeted outreach."],
    ],

    "ANALYTICS, DATA & KNOWLEDGE": [
        ["Power BI", "https://powerbi.microsoft.com",
         "Analytics & Dashboarding", "Active",
         "Data analytics and visualisation platform; builds actionable "
         "dashboards from structured datasets.",
         "Global", "Corporates; public sector; data teams",
         "Dashboards; reporting; KPIs; data connectors",
         "Structured datasets with defined schemas",
         "Microsoft / Subscription",
         "Actionable insights layer; dashboards for platform activity "
         "metrics, partnership pipeline, compliance status.",
         "Connects to multiple data sources; strong visualisation; "
         "trusted enterprise tool.",
         "Data dependent; requires clean, structured data upstream.",
         "GENE-LINK's internal analytics (partnership pipeline; compliance "
         "completion rates; geographic coverage gaps) could run on Power BI "
         "for the admin team.",
         "Insights and monetisation potential via analytics layer.",
         "Dashboard reporting; data connector ecosystem",
         "Medium–High", "None", "Medium", "Dominique",
         "[SRS-REQ] Platform admin dashboard should implement KPI "
         "monitoring (partnership pipeline, compliance completion, "
         "geographic coverage) — Power BI is a candidate tool for "
         "the internal analytics layer."],

        ["Crunchbase", "https://www.crunchbase.com",
         "Company Data Platform", "Active",
         "Database of companies, funding rounds, investors, and market "
         "intelligence. Discovery-focused.",
         "Global", "Investors; corporates; analysts",
         "Company search; funding data; investor mapping; filtering",
         "Structured company and investment profiles",
         "Private / Subscription",
         "Structured company discovery; sector and geography filtering; "
         "funding history as a commercial viability signal.",
         "Strong structured discovery; well-maintained company data.",
         "Static; no interaction layer; subscription cost.",
         "GENE-LINK's EU company onboarding could cross-reference "
         "Crunchbase to pre-populate company profiles with sector, "
         "size, and funding data.",
         "Add analytics and insights layer on top of static discovery.",
         "Filtering; company structured data; sector tagging",
         "Medium", "None", "Medium", "Dominique", ""],

        ["Dealroom", "https://dealroom.co",
         "Ecosystem Analytics", "Active",
         "Platform mapping startup ecosystems, investment trends, and "
         "innovation geography.",
         "Global", "Investors; governments; innovation agencies",
         "Ecosystem mapping; analytics; startup profiling; dashboards",
         "Structured ecosystem datasets",
         "Private / Subscription",
         "Ecosystem mapping approach — could inform GENE-LINK's "
         "geographic visualisation of bio-innovation capacity across "
         "Global South countries.",
         "Strong geographic and sectoral ecosystem insights.",
         "Limited interaction; analytics-only; subscription cost.",
         "GENE-LINK's platform analytics could adopt Dealroom's "
         "ecosystem mapping approach to visualise provider capacity "
         "by country and resource type.",
         "Add analytics and geographic mapping of bio-innovation ecosystems.",
         "Ecosystem dashboards; geographic mapping",
         "Medium", "None", "Medium", "Dominique", ""],

        ["Notion", "https://www.notion.so",
         "Knowledge & Workspace", "Active",
         "Flexible modular workspace for structured knowledge management "
         "and team collaboration.",
         "Global", "Startups; NGOs; teams",
         "Databases; documents; linked pages; flexible structure",
         "Flexible database structures; linked content",
         "Private SaaS / Subscription",
         "Clean, modular knowledge organisation applicable to GENE-LINK's "
         "internal knowledge management and prototype documentation.",
         "Clean UX; modular design; rapid prototyping of knowledge "
         "structures.",
         "Limited governance at scale; not suitable as a public-facing "
         "platform.",
         "Useful for GENE-LINK team's internal knowledge management "
         "and as a prototyping tool for platform content structure.",
         "Prototype knowledge layer before committing to full CMS.",
         "Modular knowledge design; linked database approach",
         "Medium", "None", "Medium", "Dominique", ""],

        ["Confluence", "https://www.atlassian.com/software/confluence",
         "Enterprise Knowledge Base", "Active",
         "Enterprise documentation and knowledge management with "
         "version control and permissions.",
         "Global", "Corporates; institutions",
         "Structured documentation; version control; permissions; search",
         "Hierarchical structured content + permissions",
         "Enterprise SaaS / Subscription",
         "Version-controlled documentation; institutional knowledge "
         "governance; suitable for compliance documentation management.",
         "Strong governance; version control; structured knowledge.",
         "Heavy UX; not suitable for public-facing content; "
         "Atlassian ecosystem dependency.",
         "GENE-LINK's internal compliance documentation (SOP; regulatory "
         "updates; team knowledge) should follow Confluence's structured "
         "version-controlled model.",
         "Institutional knowledge layer for compliance documentation.",
         "Version control; structured documentation; permissions",
         "Medium", "None", "Medium", "Dominique", ""],
    ],

    "REGULATORY COMPLIANCE": [
        ["ABS Clearing-House (ABSCH)", "https://absch.cbd.int",
         "Regulatory Compliance / Reference Database", "Active",
         "Global ABS compliance information system under Nagoya Protocol "
         "Article 14. Authoritative source for national ABS regulations, "
         "NFPs, CNAs, and IRCCs. The trust layer backbone for GENE-LINK.",
         "Global (196 CBD Parties)",
         "Governments; regulators; researchers; compliance teams; platforms",
         "Regulatory data search; IRCC registry; NFP/CNA directories; "
         "checkpoint records; limited API",
         "National regulatory datasets; IRCC records; permit metadata",
         "Government / Public funding (CBD Secretariat)",
         "Authoritative ABS compliance data; IRCC verification; NFP/CNA "
         "contact details; national ABS measures. Most critical data "
         "source for GENE-LINK.",
         "Authoritative trust and compliance data — GENE-LINK's value "
         "is being the UX layer on top of ABSCH, not a competitor.",
         "Poor UX; limited API documentation; data completeness varies "
         "by country; government submission cadence means some data lags.",
         "GENE-LINK wraps ABSCH data in a dramatically improved user "
         "interface. The raw ABSCH portal should never be the user's "
         "primary touchpoint.",
         "Trust layer integration: ABSCH data as compliance backbone; "
         "GENE-LINK adds actionability for non-government users.",
         "Compliance signals; regulatory data integration",
         "Very High", "High", "Very High", "Both / Merged",
         "[SRS-REQ] ABSCH API is the authoritative backend for all "
         "regulatory compliance data. Platform must never replicate or "
         "duplicate ABSCH data but must expose it through a dramatically "
         "improved UX — GENE-LINK is the UX layer on top of ABSCH."],

        ["TRACES NT (EU)", "https://food.ec.europa.eu/animals/traces_en",
         "Regulatory Compliance / Trade", "Active",
         "EU Commission platform for certification of movements of "
         "animals, animal products, food, feed, and plants across EU borders.",
         "European Union",
         "EU importers; exporters; veterinary/phytosanitary authorities",
         "Standardised movement certificates; real-time tracking; "
         "authority integration; audit trail",
         "Structured certificate records; regulatory authority links",
         "EU Government / Public funding",
         "Certificate-style structured compliance record model applicable "
         "to GENE-LINK's genetic resource shipment compliance documents.",
         "Clean certificate/permit workflow with structured data fields — "
         "the model for regulated biological material movement records.",
         "Not specific to genetic resources or ABS; no matchmaking.",
         "GENE-LINK shipment compliance records should adopt a certificate-"
         "style structured data model similar to TRACES.",
         "Add genetic resource and ABS specificity to the certificate model.",
         "Structured certificate model; regulated movement records",
         "Low", "Low", "Medium", "Platform Team",
         "[SRS-REQ] GENE-LINK shipment compliance records must adopt a "
         "certificate-style structured data model with mandatory fields, "
         "time-stamping, and audit trail, modelled on TRACES NT."],

        ["CITES Trade Portal", "https://trade.cites.org",
         "Regulatory Compliance / Trade Data", "Active",
         "Official portal for CITES trade data; tracks international "
         "trade in specimens of species listed under CITES Appendices.",
         "Global (CITES Parties)",
         "Researchers; enforcement authorities; compliance teams",
         "Species trade records; permit data; country-pair analysis; "
         "downloadable datasets",
         "Species-level trade records by country pair",
         "Government / Public funding (CITES Secretariat)",
         "Species-level compliance data store at scale; downloadable "
         "datasets; species-level permit tracking.",
         "Transparent, species-level compliance data at scale — model "
         "for what a species compliance data store looks like.",
         "CITES-listed species only; not ABS-specific; no matchmaking.",
         "GENE-LINK should flag CITES listing status in species search "
         "results as an additional compliance layer alongside ABS.",
         "Complement ABS compliance with CITES trade permit screening.",
         "Species-level permit records; compliance transparency",
         "High", "Medium", "Medium", "Platform Team",
         "[SRS-REQ] Species search must cross-check CITES appendix status "
         "and flag any additional trade permit requirements before "
         "initiating the ABS compliance workflow."],
    ],

    "INNOVATION NETWORKS": [
        ["Global Innovation Exchange",
         "https://www.globalinnovationexchange.org",
         "Innovation Platform / Discovery", "Active",
         "Platform showcasing development innovations and connecting "
         "stakeholders; USAID-supported.",
         "Global", "NGOs; donors; innovators",
         "Listings; discovery browsing; innovation profiles",
         "Project-based innovation profiles",
         "Donor supported / Grant funded",
         "Strong discovery layer for development innovations; "
         "portfolio visibility model.",
         "Strong discovery and portfolio visibility; multi-stakeholder reach.",
         "Limited transaction layer; weak communication and follow-through.",
         "GENE-LINK should add collaboration and deal-facilitation "
         "layers that GIE lacks.",
         "Add ABS-specific collaboration layer on top of discovery.",
         "Portfolio visibility; innovation listing model",
         "Medium–High", "None", "Medium", "Dominique", ""],

        ["CTCN (Climate Technology Centre)",
         "https://www.ctc-n.org",
         "Innovation Network (Request-Response)", "Active",
         "UN-led climate technology collaboration network with request-"
         "driven matching model.",
         "Global", "Governments; institutions",
         "Technology requests; matchmaking; facilitated responses",
         "Country and project profiles; request-response system",
         "UN-led / Donor funded",
         "Request-driven matching model (countries submit requests, "
         "network matches with providers) — applicable to GENE-LINK's "
         "EU company resource request workflow.",
         "Strong network orchestration; government legitimacy; "
         "request-driven matching is highly user-centric.",
         "Not a marketplace; no transaction layer; slow facilitation.",
         "GENE-LINK's request-driven matching (EU company posts a "
         "resource need; platform matches with Southern providers) "
         "should follow CTCN's request-response model.",
         "Add structured marketplace layer to the network model.",
         "Request-response matching; network orchestration",
         "Medium", "None", "Medium", "Dominique",
         "[SRS-REQ] Platform must support both supply-driven (Southern "
         "provider lists a resource) and demand-driven (EU company posts "
         "a resource need) matching flows, following CTCN's request-"
         "response model."],
    ],

    "CAPACITY DEVELOPMENT": [
        ["Nagoya Protocol Hub",
         "https://nagoyaprotocol-hub.de",
         "Capacity Development / ABS Guidance", "Active",
         "German-government-supported hub with ABS training materials, "
         "country-specific Easy ABS Guides, and e-learning resources.",
         "Global (German/EU focus)",
         "Researchers; companies; practitioners",
         "Country ABS guides; step-by-step compliance walkthroughs; "
         "training modules; newsletter",
         "PDF guides; web articles; structured country profiles",
         "Free / German government funding",
         "Plain-language, country-specific ABS compliance guides; "
         "step-by-step user journeys; practitioner-level guidance.",
         "Country-specific guides are the most user-friendly ABS "
         "compliance content available — essential reference for "
         "GENE-LINK's guided walkthrough design.",
         "Static content; no platform interactivity; no data integration; "
         "updates can lag regulatory changes.",
         "GENE-LINK should embed Nagoya Protocol Hub country guides "
         "within the compliance module, with dynamic content drawn "
         "from the ABSCH API to keep it current.",
         "Complement static Hub content with dynamic ABSCH-backed data.",
         "Plain-language country walkthroughs; practitioner guidance",
         "High", "High", "High", "Platform Team",
         "[SRS-REQ] When a user selects a provider country, platform "
         "must surface the relevant Nagoya Protocol Hub country guide "
         "alongside live ABSCH regulatory data."],

        ["BioInnovation Africa (BIA)",
         "https://www.abs-biotrade.info",
         "Capacity Development / ABS Matchmaking", "Active",
         "GIZ-managed platform documenting ABS best practices, value chain "
         "maps, MAT negotiation training, and African-European partnership "
         "case studies. 7,400+ field operators.",
         "Africa–Europe", "African bio-innovators; EU companies; practitioners",
         "ABS manuals; MAT training packages; value chain case studies; "
         "field operator network",
         "PDF manuals; case study documents; partner contact details",
         "Free / GIZ / EC project funding",
         "Practitioner-level ABS guidance for African contexts; MAT "
         "negotiation training; value chain maps; case studies.",
         "Highest-quality practitioner content for Africa; real partnership "
         "case studies; GIZ institutional backing.",
         "Africa-only; no interactive features; no live data; "
         "no Asian or Latin American coverage.",
         "GENE-LINK guidance module for African provider countries must "
         "integrate BIA materials. GIZ partnership should be formalised "
         "for content updates.",
         "Add interactivity, global coverage, and live data on top of "
         "BIA's strong content foundation.",
         "ABS manuals; MAT training; value chain documentation",
         "High", "High", "High", "Platform Team",
         "[SRS-REQ] Platform guidance module for African provider countries "
         "must integrate BIA ABS manuals and MAT training materials via "
         "a formal GIZ content partnership."],

        ["UEBT Training Platform",
         "https://www.ethicalbiotrade.org/training",
         "Capacity Development (Industry)", "Active",
         "Union for Ethical BioTrade training and certification programmes "
         "for EU companies sourcing biodiversity-based ingredients.",
         "Global (EU company focus)",
         "Cosmetics; food; personal care companies; sourcing managers",
         "Ethical sourcing standards; company certification; "
         "training modules; supplier audit frameworks",
         "Course-based profiles; certification records",
         "Paid / UEBT membership-based",
         "EU company-facing compliance training; UEBT certification as "
         "a recognised trust signal; sector-specific (cosmetics/food).",
         "Best available industry-facing ABS/ethical sourcing training "
         "for EU cosmetics and food sector companies.",
         "Certification-focused; no matchmaking; no Southern institution "
         "directory; paid membership required.",
         "GENE-LINK company onboarding should reference UEBT standards "
         "and offer UEBT certification as a recommended trust pathway.",
         "Add matchmaking and Southern institution access on top of UEBT "
         "compliance training.",
         "Ethical sourcing certification; corporate training framework",
         "High", "Medium", "Medium", "Platform Team",
         "[SRS-REQ] Platform EU company profiles must include a field for "
         "UEBT certification status as a verified trust indicator."],

        ["Moodle", "https://moodle.org",
         "Learning Platform (Open Source)", "Active",
         "Open-source learning management system with structured courses, "
         "assessments, and certification.",
         "Global", "Education institutions; NGOs; governments",
         "Structured courses; assessments; forums; certification",
         "Course-based user profiles; enrolment records",
         "Open source / Hosting / Subscription",
         "SCORM-compatible structured learning pathways; open source; "
         "self-hostable; certificate generation.",
         "Scalable, proven LMS; SCORM compatibility allows existing CBD "
         "and WIPO courses to be embedded.",
         "Low engagement if courses are poorly designed; "
         "requires significant content investment.",
         "GENE-LINK's e-learning module could run on a Moodle instance, "
         "allowing embedding of existing SCORM courses from CBD, WIPO, "
         "and FAO without rebuilding them.",
         "Capacity building layer with certification for partner readiness.",
         "SCORM-compatible learning pathways; certification; open source",
         "High", "None", "Medium", "Dominique",
         "[SRS-REQ] Platform e-learning module should be SCORM-compatible "
         "(using Moodle or equivalent) to allow embedding of existing "
         "accredited ABS courses from CBD, WIPO, and FAO."],
    ],

    "E-LEARNING": [
        ["CBD Learning Platform",
         "https://www.cbd.int/elearning",
         "E-Learning", "Active",
         "SCORM-compliant self-paced e-learning courses from the CBD "
         "Secretariat covering the Nagoya Protocol, CBD, and related "
         "instruments. Multilingual. Certificates of completion issued.",
         "Global", "Government officials; practitioners; researchers",
         "Self-paced courses; assessments; certificates; multilingual",
         "Course-based profiles with completion records",
         "Free / CBD Secretariat public funding",
         "Best available structured ABS e-learning; multilingual; "
         "certificates of completion; SCORM-compatible for embedding.",
         "Authoritative; free; multilingual; SCORM; certificates.",
         "Government/practitioner audience; not tailored to private "
         "sector or technical users.",
         "GENE-LINK should link to CBD courses at onboarding and at "
         "relevant compliance steps rather than rebuilding the content.",
         "Complement with private-sector and Southern institution "
         "framing around the core CBD content.",
         "SCORM courses; multilingual; completion certificates",
         "High", "High", "High", "Platform Team",
         "[SRS-REQ] Platform must surface the CBD Nagoya Protocol "
         "e-learning course at user onboarding and at relevant "
         "compliance workflow steps."],
    ],

    "UX & BEHAVIOURAL REFERENCE": [
        ["Tinder", "https://tinder.com",
         "UX / Behavioural Reference", "Active",
         "Swipe-based matching enabling rapid discovery with minimal "
         "cognitive friction. Referenced for UX pattern inspiration only.",
         "Global", "General public",
         "Swipe matching; instant discovery; real-time chat",
         "Lightweight profiles with images",
         "Private platform / Subscription + ads",
         "Extreme simplicity in discovery; engagement loops; "
         "minimal friction from profile to first interaction.",
         "Elimination of friction from intent to first contact is the "
         "core insight: reduce steps between 'interested' and 'connected'.",
         "Shallow trust model; inappropriate depth for ABS partnerships.",
         "GENE-LINK's discovery UX should minimise steps from intent "
         "to first connection — not replicate swipe mechanics but "
         "adopt the principle of low-friction discovery.",
         "Inspire low-friction discovery UX; reduce time-to-first-contact.",
         "Swipe/quick-interest logic; minimal friction discovery",
         "Low–Medium", "None", "High", "Dominique",
         "[SRS-REQ] Platform discovery UX must minimise steps from "
         "intent to first contact — adopt swipe-logic principle of "
         "low cognitive friction while maintaining the trust depth "
         "appropriate for ABS partnerships."],

        ["Hinge", "https://hinge.co",
         "UX / Behavioural Reference", "Active",
         "Richer-profile matching platform focusing on deeper connection "
         "through narrative prompts and structured profile sections.",
         "Global", "General public",
         "Matching; prompt-based profiles; richer narrative sections",
         "Richer narrative profiles with prompted sections",
         "Private platform / Subscription",
         "Prompt-based profile sections elicit richer, structured "
         "information that improves match quality over minimal "
         "free-text profiles.",
         "Deeper profile-driven trust building through structured prompts.",
         "Still consumer-oriented; shallow for institutional context.",
         "GENE-LINK profiles should use structured prompt-based sections "
         "('What resource types are you looking for?', 'What does fair "
         "benefit-sharing mean to your organisation?') rather than "
         "unstructured free text.",
         "Model deeper, trust-building profiles for institutional context.",
         "Prompt-based matching; narrative profile sections",
         "Low–Medium", "None", "High", "Dominique",
         "[SRS-REQ] Platform institution profiles must include structured "
         "prompt-based sections (modelled on Hinge) that elicit richer, "
         "comparable information than unstructured free text — enabling "
         "better algorithmic matching and building trust with profile readers."],

        ["Bumble", "https://bumble.com",
         "UX / Behavioural Reference", "Active",
         "Platform controlling who initiates interaction; reduces "
         "unsolicited outreach.",
         "Global", "General public",
         "Matching; controlled-initiation messaging",
         "Profiles with interaction control rules",
         "Private platform / Subscription",
         "Controlled interaction flow: one party initiates, the other "
         "must respond within a time window.",
         "Initiation control prevents unsolicited outreach and spam.",
         "Consumer-oriented; binary initiation rule may be too rigid.",
         "GENE-LINK should consider who initiates contact (EU company "
         "or Southern provider) and whether platform rules should "
         "govern initiation to prevent power imbalances.",
         "Safer, more equitable engagement flows for North-South context.",
         "Initiation rules; controlled first-contact logic",
         "Low", "None", "Medium", "Dominique",
         "[SRS-REQ] Platform contact initiation model should be designed "
         "to prevent power-imbalance dynamics: consider whether Southern "
         "providers should have the ability to accept or decline contact "
         "before EU users can initiate direct communication."],
    ],
}


# ---------------------------------------------------------------------------
# Sheet I-B — Architecture Notes (Dominique's data, all 29 rows preserved)
# ---------------------------------------------------------------------------

ARCH_COLS = [
    "Platform / Source", "URL",
    "Architecture / Technical Pattern",
    "Onboarding / Registration Flow",
    "Profile / Data Model Structure",
    "Matching / Discovery Algorithm",
    "Search & Filter Capabilities",
    "Notification / Communication Features",
    "API / Integration Capabilities",
    "Mobile Responsiveness",
    "Multilingual Support",
    "Governance / Admin Model",
    "Key Design Feature to Adopt",
    "Key Design Failure to Avoid",
    "Relevance to GENE-LINK\n(High / Med / Low)",
    "Added By", "Notes",
]

# Raw rows from Dominique (29 entries) — columns match ARCH_COLS
ARCH_ROWS = [
    ["b2match", "https://www.b2match.com", "Structured B2B matchmaking with scheduling", "Form-based with collaboration interests", "Structured: offers/requests/expertise/org-type", "Rule-based + user-driven", "Strong: sector, role, collaboration type",
        "Meeting requests; confirmations; reminders", "Moderate (event integrations)", "Moderate", "Strong (EU)", "Centralised (event owner)", "Structured collaboration fields", "Rigid UX; limited engagement outside events", "High", "Dominique", "Strong EU baseline"],
    ["Grip", "https://www.grip.events", "AI-driven matchmaking over behavioural data", "Guided onboarding with tagging + interests", "Roles, interests, behavioural signals", "AI recommendation + interest + behaviour", "Advanced filtering + recommendations",
        "Match suggestions; meeting prompts", "Strong (event + CRM integrations)", "Strong", "Moderate", "Centralised (event organisers)", "Mutual interest matching logic", "Event-only engagement lifecycle", "High", "Dominique", "Best-in-class UX benchmark"],
    ["Swapcard", "https://www.swapcard.com", "Hybrid: content + networking + matchmaking", "Event-driven onboarding via participation", "Rich profiles + engagement data", "AI + activity-based recommendations", "Moderate",
        "Content alerts; chat; notifications", "Strong (event ecosystem)", "Strong", "Moderate", "Centralised", "Content + networking integration", "Episodic event-bound engagement", "Medium–High", "Dominique", "Front-end UX reference"],
    ["Discourse", "https://www.discourse.org", "Forum-based knowledge and discussion", "Simple signup; optional structured onboarding", "User profiles linked to contributions + topics",
        "Topic-based (not algorithmic)", "Strong tagging and search", "Thread notifications; replies; mentions", "Strong (API + plugins)", "Strong", "Strong", "Moderated community", "Persistent searchable knowledge base", "Low engagement without facilitation", "High", "Dominique", "Knowledge layer anchor"],
    ["Microsoft Teams", "https://www.microsoft.com/teams", "Enterprise comms and collaboration", "Account-based (org-linked)", "User identity tied to org + permissions", "No matchmaking", "Basic search across chats/files", "Messaging; calls; file sharing; alerts",
     "Strong (Microsoft ecosystem)", "Strong", "Strong", "Enterprise (IT-controlled)", "Secure trusted communication layer", "Fragmented across channels", "High", "Dominique", "Primary secure comms option"],
    ["DocSend", "https://www.docsend.com", "Secure document sharing with access control", "Simple link-based",
        "Document-centric (no user profiles)", "Document access alerts + analytics", "Limited", "Strong", "Low", "N/A", "Owner-controlled (doc-level)", "Controlled visibility + tracking", "Lack of collaborative workflow", "High", "Dominique", "Bridge to deal-making"],
    ["Intralinks", "https://www.intralinks.com", "Enterprise virtual data room", "Formal onboarding with permissions", "Document-centric with strict access control", "N/A", "Secure collaboration notifications",
        "Limited", "Moderate", "Moderate", "Highly controlled enterprise governance", "High-security document sharing", "Over-complex for early-stage", "Medium–High", "Dominique", "Later-stage use"],
    ["Fiverr", "https://www.fiverr.com", "Marketplace with reputation + transaction", "Account + service listing", "Service profiles with ratings + reviews",
        "Search + ranking (relevance + reputation)", "Strong: service, price, rating", "Messaging; order updates; alerts", "Limited (internal ecosystem)", "Strong", "Moderate", "Platform-controlled", "Reputation accumulation within platform", "Users bypassing platform", "High", "Dominique", "Retention model benchmark"],
    ["Upwork", "https://www.upwork.com", "Marketplace with contracts + escrow", "Account + proposal-based", "Profiles with work history; ratings; skills", "Bidding + ranking algorithm", "Strong filtering",
        "Messaging; milestones; payments", "Moderate", "Strong", "Moderate", "Platform-controlled", "Escrow-based trust mechanism", "Complex onboarding", "High", "Dominique", "Structured contracting model"],
    ["LinkedIn", "https://www.linkedin.com", "Professional identity + network", "Simple with progressive profile completion",
        "Professional identity (roles; experience; network)", "Connection-based + algorithmic feed", "Moderate", "Messaging; feed notifications; alerts", "Strong (APIs; integrations)", "Strong", "Strong", "Centralised (platform)", "Identity + credibility signals", "Weak transaction capability", "High", "Dominique", "Profile benchmark"],
    ["Rocket.Chat", "https://www.rocket.chat", "Self-hosted secure messaging", "Account-based", "User accounts + channels", "Basic search", "Real-time messaging",
        "Self-hosted API (flexible)", "Moderate", "Moderate", "Self-hosted (org-controlled)", "Data sovereignty + control", "Less polished UX", "Medium", "Dominique", "Secure alternative"],
    ["Moodle", "https://moodle.org", "Learning management system", "Course-based enrolment", "Profiles linked to course participation", "Course search", "Forum; course notifications",
        "Strong", "Moderate", "Strong", "Admin-controlled (institutional)", "Structured learning pathways", "Low engagement if poorly designed", "Medium", "Dominique", "Capacity building layer"],
    ["Notion", "https://www.notion.so", "Flexible modular knowledge system", "Open-ended onboarding", "Flexible database structures", "Search across content", "Comments; mentions", "Strong",
        "Strong", "Moderate", "Flexible (workspace-level)", "Clean modular knowledge organisation", "Limited governance at scale", "Medium", "Dominique", "Prototype knowledge layer"],
    ["Confluence", "https://www.atlassian.com/software/confluence", "Enterprise documentation platform", "Account-based", "Structured documentation with permissions", "Strong search",
        "Comments; collaboration tools", "Strong", "Moderate", "Strong", "Enterprise governance", "Version control + structured knowledge", "Heavy UX", "Medium", "Dominique", "Institutional knowledge base"],
    ["Power BI", "https://powerbi.microsoft.com", "Analytics and dashboard platform", "Data integration onboarding", "Structured datasets", "N/A — dashboard alerts",
        "Strong", "Moderate", "Moderate", "Governed (enterprise)", "Actionable insights", "Data dependency", "Medium–High", "Dominique", "Insights layer"],
    ["ABS Clearing-House", "https://absch.cbd.int", "Centralised compliance database", "Government data submission", "Regulatory datasets", "Searchable database", "Minimal",
        "Government integrations", "Low", "Strong", "Government-led", "Authoritative compliance data", "Poor UX and usability", "Very High", "Dominique", "Critical trust layer"],
    ["WIPO GREEN", "https://www.wipo.int/wipogreen", "Technology listing + matchmaking database", "Profile submission", "Technology and need profiles", "Self-directed discovery",
        "Basic search", "Limited messaging", "Moderate", "Moderate", "Moderate", "Moderated platform", "Relevant thematic model", "Low engagement", "Very High", "Dominique", "Key analogue"],
    ["Global Innovation Exchange", "https://www.globalinnovationexchange.org", "Innovation listing platform", "Profile/project submission", "Project-based profiles", "Browse-based discovery",
        "Moderate search", "Limited communication", "Moderate", "Moderate", "Moderate", "Donor-supported", "Strong discovery layer", "Weak transaction layer", "Medium", "Dominique", "Discovery reference"],
    ["CTCN", "https://www.ctc-n.org", "Networked request-response platform", "Request submission", "Country/project profiles", "Request-driven matching", "Moderate search",
        "Facilitated communication", "Moderate", "Moderate", "Moderate", "UN-led", "Strong network orchestration", "Not a marketplace", "Medium", "Dominique", "System-level reference"],
    ["Tinder", "https://tinder.com", "Swipe-based matching", "Fast with minimal data", "Lightweight profiles",
        "Mutual interest (swipe) algorithm", "Limited", "Instant notifications; chat", "Limited", "Strong", "Strong", "Centralised", "Extreme simplicity in discovery", "Shallow trust model", "High", "Dominique", "Discovery UX"],
    ["Hinge", "https://hinge.co", "Profile-rich matching with prompts", "Guided with prompted sections", "Richer narrative profiles", "Compatibility-based matching", "Moderate",
        "Notifications; chat", "Limited", "Strong", "Moderate", "Centralised", "Deeper profile-driven trust", "Still consumer-oriented", "High", "Dominique", "Trust UX"],
    ["Bumble", "https://bumble.com", "Controlled interaction platform", "Simple onboarding", "Profiles with preferences", "Matching + controlled initiation", "Moderate",
        "Chat with initiation rules", "Limited", "Strong", "Moderate", "Centralised", "Controlled interaction flow", "Limited B2B relevance", "Medium", "Dominique", "Interaction logic"],
    ["Airbnb", "https://www.airbnb.com", "Trust-based marketplace", "Account + identity verification", "Listings + reviews + profiles", "Search + ranking", "Strong filtering",
        "Booking; messaging; alerts", "Strong", "Strong", "Strong", "Platform-governed", "Trust signals (reviews; identity)", "Platform dependency", "Very High", "Dominique", "Trust benchmark"],
    ["AngelList", "https://angel.co", "Startup-investor marketplace", "Account + startup listing", "Startup profiles + funding data", "Discovery + investor matching", "Moderate",
        "Messaging; deal updates", "Moderate", "Strong", "Moderate", "Platform-controlled", "Investment matchmaking", "Limited scale outside niche", "Medium", "Dominique", "Investment logic"],
    ["Crunchbase", "https://www.crunchbase.com", "Data discovery platform", "Account-based", "Structured company data", "Search-driven", "Strong", "Limited notifications",
        "Strong", "Strong", "Moderate", "Platform-controlled", "Structured discovery data", "No interaction layer", "Medium", "Dominique", "Data layer"],
    ["Dealroom", "https://dealroom.co", "Ecosystem analytics platform", "Account-based", "Structured datasets", "Analytics-driven discovery", "Strong",
        "Dashboard alerts", "Strong", "Strong", "Moderate", "Platform-controlled", "Ecosystem insights", "Limited interaction", "Medium", "Dominique", "Insights layer"],
    ["ResearchGate", "https://www.researchgate.net", "Academic network", "Simple onboarding", "Researcher profiles + publications", "Network + topic-based", "Moderate",
        "Notifications; messaging", "Limited", "Strong", "Moderate", "Platform-controlled", "Academic credibility", "Weak private sector integration", "Medium", "Dominique", "Research layer"],
    ["CORDIS", "https://cordis.europa.eu", "EU research database", "Project-based profiles", "Searchable database", "Minimal", "N/A", "EU systems",
        "Moderate", "Strong", "EU-managed", "Strong EU legitimacy", "Static and non-interactive", "High", "Dominique", "EU integration"],
    ["GitHub", "https://github.com", "Collaborative development platform", "Account-based", "Code repositories + contributors", "Searchable repositories", "Notifications; collaboration tools",
        "Strong", "Strong", "Moderate", "Platform + open source governance", "Transparent collaboration", "Versioning complexity", "Medium", "Dominique", "Collaboration logic"],
]


# ---------------------------------------------------------------------------
# Data source sheets: shared 18-column schema
# ---------------------------------------------------------------------------

DS_COLS = [
    "Source Name", "Description", "Key Data Available",
    "Geographic Coverage",
    "Access Type\n(Free / Paid / Restricted)",
    "Technical Access\n(API / Portal / Download / Manual)",
    "URL / Endpoint", "Data Format & Standards",
    "Update Frequency", "Data Quality / Completeness *",
    "Authentication Required", "Cost Indicator",
    "Integration Complexity\n(Low / Med / High)",
    "Primary Platform Function",
    "Cross-Scan Linkage", "Priority\n(High / Med / Low)",
    "Notes & Limitations", "SRS Requirement",
]

# --- II-A: Regulatory ---
REG_ROWS = [
    ["ABS Clearing-House (ABSCH)",
     "Central information-sharing platform under Nagoya Protocol Article 14. "
     "Authoritative source for national ABS regulations, NFPs, CNAs, and IRCCs.",
     "NFPs and CNAs per country; IRCC registry; checkpoint records; national "
     "ABS measures; permit templates",
     "Global (196 CBD Parties)", "Free", "Public API",
     "https://absch.cbd.int/api/v2016",
     "JSON / REST. CBD-defined schema.",
     "Continuous (as submitted by Parties)",
     "Variable by country — major providers generally complete; some nations sparse",
     "No (read-only endpoints open)", "None", "Low",
     "Compliance / Due Diligence", "Governance Scan", "High",
     "IRCC issuance varies by country (Brazil issues SisGen receipts instead of "
     "formal IRCCs). Rate limits apply.",
     "[SRS-REQ] Platform must query ABSCH API to auto-populate NFP/CNA details "
     "and IRCC verification when a provider country is selected."],

    ["CGHSS AbsMaster Dataset",
     "Country-level ABS governance dataset (Georgetown / CBD study, BMJ Public "
     "Health doi:10.1136/bmjph-2024-001800). Machine-readable CSV derived from "
     "ampeid.org and ABSCH.",
     "CBD ratification status; Nagoya ratification dates; ABS legislation "
     "indicators; compliance portal existence per country",
     "Global (all CBD Parties)", "Free (open data)", "Public Download (GitHub CSV)",
     "https://github.com/cghss/ABS/blob/main/data/AbsMaster.csv",
     "CSV. Flat-file, manually curated.", "Periodic (research-driven)",
     "High for CBD/Nagoya ratification; moderate for operational status fields",
     "No", "None", "Low", "Compliance / Search", "Governance Scan", "High",
     "Best machine-readable summary of global ABS ratification landscape. "
     "Should be ingested to pre-populate regulatory status for all 15 focus "
     "countries.",
     "[SRS-REQ] Platform regulatory status fields for all provider countries "
     "must be pre-populated from AbsMaster dataset on launch."],

    ["SisGen (Brazil)",
     "Brazil's digital ABS registration and compliance management system. "
     "Mandatory for any activity involving Brazilian genetic heritage or "
     "associated TK.",
     "Access registrations; shipment registrations; R&D activity records; "
     "benefit-sharing agreement records",
     "Brazil only", "Free", "Portal (no public API)",
     "https://sisgen.gov.br", "Web portal only.", "Continuous (per-registration)",
     "Authoritative for Brazil — it is the legal record system",
     "Yes (Brazilian partner institution account required for foreign users)",
     "None", "High",
     "Compliance / Due Diligence", "Governance Scan", "High",
     "Foreign users cannot register directly — mandatory MoU with Brazilian "
     "institution required. Platform can guide but not automate registration.",
     "[SRS-REQ] Platform must provide a guided SisGen registration checklist "
     "for Brazil-origin resources and flag the mandatory partner institution "
     "requirement."],

    ["NBA ABS e-Filing Portal (India)",
     "India's online portal for National Biodiversity Authority approvals. "
     "Handles access, commercial use, IPR filings, and Certificate of Origin "
     "for cultivated medicinal plants.",
     "Form I (access); Form III (IPR); CoO for cultivated plants; "
     "application tracking",
     "India only", "Free", "Portal (no public API)",
     "https://absefiling.nic.in", "Web portal only.", "Continuous (per-application)",
     "Authoritative for India",
     "Yes (applicant account required)", "None", "High",
     "Compliance / Due Diligence", "Governance Scan", "High",
     "Different forms for different activity types. 2023 amendments added "
     "Certificate of Origin exemption pathway for cultivated medicinal plants.",
     "[SRS-REQ] Platform compliance module for India must implement activity-"
     "type branching to surface the correct NBA form."],

    ["EU ABS Regulation 511/2014",
     "Official EU guidance on due diligence obligations for EU-based users "
     "of genetic resources under the Nagoya Protocol.",
     "Due diligence declaration requirements; IRCC verification steps; "
     "checkpoint reporting obligations; sector-specific guidance",
     "European Union (+ UK separately)", "Free", "Documentation / Portal",
     "https://environment.ec.europa.eu/topics/nature-and-biodiversity/"
     "access-and-benefit-sharing_en",
     "PDF / HTML guidance documents", "As regulation changes",
     "Authoritative for EU user obligations", "No", "None", "Low",
     "Compliance", "Governance Scan", "High",
     "UK OPSS/DEFRA guidance applies separately to UK entities post-Brexit. "
     "Platform must distinguish EU vs. UK jurisdiction.",
     "[SRS-REQ] Platform onboarding for EU company users must surface the "
     "correct due diligence obligations (EU 511/2014 or UK equivalent) based "
     "on company registration country."],

    ["WIPO Patentscope", "https://patentscope.wipo.int",
     "Global patent database covering biotech and pharmaceutical patents. "
     "Broader international coverage than Espacenet (EPO); includes PCT "
     "applications. Host: WIPO.",
     "Patent publications; IPC codes; applicant data; sequence listings; "
     "biotech-specific search; prior art; country filing data",
     "Global", "Free", "Public API",
     "https://patentscope.wipo.int/search/en/help/data_download.jsf",
     "XML / JSON via WIPO API", "Continuous (weekly)",
     "High — authoritative international patent record",
     "Yes (WIPO API key — free registration)", "None", "Low",
     "Due Diligence", "Company Scan / Governance Scan", "Very High",
     "Complementary to Espacenet: use WIPO Patentscope for PCT applications "
     "and developing-country patents; Espacenet for European patents. "
     "Critical for biopiracy detection.",
     "[SRS-REQ] Platform due diligence must query WIPO Patentscope "
     "alongside Espacenet for comprehensive biopiracy screening of target "
     "genetic resources."],

    ["Lens.org",
     "Open patent and scholarly database linking patents to research "
     "publications. 200M+ records. Strong API. Bridges science and innovation.",
     "Patent metadata; scholarly citations; applicant/inventor networks; "
     "patent families; open access to full dataset",
     "Global", "Free (open access)", "Public API (strong)",
     "https://api.lens.org",
     "JSON / REST. Linked open data.", "Ongoing",
     "Very high — peer-reviewed patent + scholarly data linkage",
     "Yes (free API key)", "None", "Low",
     "Due Diligence / Search", "Company Scan", "Very High",
     "Unique in linking patents directly to the scientific papers that "
     "cite them — essential for tracking genetic resource utilisation "
     "from publication to patent.",
     "[SRS-REQ] Lens.org should be the primary open-access source for "
     "patent screening in the due diligence workflow, supplementing "
     "WIPO Patentscope."],
]

# --- II-B: Company & Innovation ---
COMP_ROWS = [
    ["John Snow Labs Biotech Pharma Directory",
     "Curated directory of US and EU biotech and pharmaceutical companies "
     "with specialisation, contacts, facility registrations. "
     "Available via Databricks Marketplace.",
     "Company name; sector; specialisation; contact details; facility type; "
     "geographic presence; regulatory registrations",
     "US and EU (primary)", "Paid", "Databricks API",
     "https://marketplace.databricks.com (John Snow Labs listing)",
     "Structured JSON via Databricks connector", "Periodic updates",
     "High for registered pharma/biotech; lower for smaller SMEs",
     "Yes (Databricks account + licence)",
     "Subscription (Databricks Marketplace pricing)", "High",
     "Matchmaking", "Company Scan", "High",
     "Best available structured database for EU/US biotech company profiling. "
     "Company Scan team should coordinate on licence sharing.",
     "[SRS-REQ] Platform EU company profiles should draw from this database "
     "for pre-population; users should be able to claim and enrich profiles."],

    ["EBSCO Biotechnology Source",
     "Paywalled full-text research database for life sciences and pharma R&D. "
     "7,800+ journals, industry publications, conference proceedings.",
     "Full-text journals; industry market reports; R&D pipeline intelligence; "
     "company profiles; regulatory science",
     "Global (EU/US heavy)", "Paid (institutional subscription)",
     "Authenticated API (institutional)",
     "https://about.ebsco.com/products/research-databases/biotechnology-source",
     "Proprietary EBSCO EDS API", "Continuous",
     "Very high — peer-reviewed and industry-validated",
     "Yes (institutional subscription)", "Significant institutional fee", "High",
     "Search", "Company Scan", "Medium",
     "Platform should provide gateway links rather than direct integration. "
     "Most target users will have institutional access.",
     "[SRS-REQ] Platform search should surface EBSCO links for R&D market "
     "intelligence with clear subscription flag and open-access fallbacks."],

    ["UEBT Member Directory",
     "Directory of Union for Ethical BioTrade member companies committed to "
     "ethical biodiversity sourcing. EU cosmetics, food, and personal care focus.",
     "Member company names; sectors; geographic presence; certification status",
     "Global (EU focus)", "Restricted (UEBT members/partners)",
     "Portal / Manual",
     "https://www.ethicalbiotrade.org/members",
     "HTML (no API)", "As updated",
     "High for UEBT-certified companies; limited universe",
     "Yes (UEBT partnership required)",
     "Requires UEBT partnership/membership", "High",
     "Matchmaking", "Company Scan", "High",
     "UEBT certification is a high-value trust signal. Formal data-sharing "
     "agreement with UEBT is strategically important for GENE-LINK.",
     "[SRS-REQ] Platform should recognise UEBT certification as a verified "
     "trust indicator on EU company profiles; pursue data partnership."],

    ["CORDIS (EU Research Database)",
     "EU platform for Horizon Europe and Framework Programme funded research "
     "results; comprehensive database of EU research projects and partners.",
     "Project profiles; funding data; consortium member details; results; "
     "publication links",
     "EU", "Free", "Limited API",
     "https://cordis.europa.eu",
     "Structured project data; REST API available", "Ongoing",
     "High for EU-funded projects; does not cover private R&D",
     "No", "None", "Medium",
     "Search / Matchmaking", "Company Scan", "High",
     "Identifies EU research teams and companies active in bio-innovation "
     "projects — a warm prospecting list for EU user onboarding.",
     "[SRS-REQ] Platform EU company onboarding should cross-reference "
     "CORDIS to identify companies active in relevant EU-funded research."],

    ["Espacenet (EPO Patents)",
     "European Patent Office's free patent search database. Critical for "
     "biopiracy screening and identifying EU companies active in "
     "bioprospecting-related patents.",
     "Patent publications; IPC codes (A61K pharma; A01N biocides; C12N biotech); "
     "applicant names; sequence listings; prior art",
     "Global (EPO + national offices)", "Free", "Public API (OPS)",
     "https://ops.epo.org",
     "XML / JSON via OPS REST API", "Weekly updates",
     "High — authoritative European patent record",
     "Yes (OPS API key — free registration)", "None", "Low",
     "Due Diligence / Search", "Company Scan / Governance Scan", "High",
     "Use for European patents specifically; complement with WIPO Patentscope "
     "for broader international coverage.",
     "[SRS-REQ] Patent screening in due diligence workflow must query "
     "Espacenet for European patents; results framed as advisory flags "
     "requiring legal review."],

    ["OpenCorporates",
     "World's largest open database of companies aggregating data from "
     "official government company registries globally.",
     "Company registration; registered address; status (active/dissolved); "
     "director information; registry source",
     "Global (80+ jurisdictions)", "Free (basic) / Paid (advanced)",
     "Public API (basic) / Auth API (advanced)",
     "https://api.opencorporates.com",
     "JSON / REST", "Regular (synced from registries)",
     "High for open-registry jurisdictions; variable elsewhere",
     "No (basic) / Yes (API key for advanced)",
     "None (basic); API pricing for bulk", "Low",
     "Due Diligence", "Company Scan", "Medium",
     "Useful for verifying company registration and legal standing before "
     "executing benefit-sharing agreements.",
     "[SRS-REQ] Platform due diligence must include an OpenCorporates "
     "company registration verification step before MTA execution."],

    ["Proposed: GENE-LINK Opportunity Pipeline",
     "Custom database to be built as a core platform asset. Tracks and "
     "manages the full partnership opportunity pipeline from initial interest "
     "to executed agreement.",
     "Partnership opportunities; EU company interest signals; provider "
     "resource offers; deal stage; ABS compliance status; match score",
     "Global", "Platform-internal (private)", "Platform database",
     "Proposed — to build", "Custom schema (GGBN-aligned)",
     "Real-time", "High (platform-owned data)",
     "N/A", "N/A", "High (to build)",
     "Matchmaking / Compliance", "All scans", "Very High",
     "This is the core matchmaking engine dataset. Must be designed before "
     "platform development begins.",
     "[SRS-REQ] Platform must maintain an opportunity pipeline database "
     "tracking deal stage, ABS compliance status, and match scores for "
     "all active provider-user interactions."],
]

# --- II-C: Resource & Biodiversity ---
BIO_ROWS = [
    ["GBIF (Global Biodiversity Information Facility)",
     "World's largest open-access biodiversity data infrastructure. "
     "1.5B+ occurrence records. Primary tool for provider-country mapping.",
     "Species occurrence records; taxonomic backbone; country checklists; "
     "dataset provenance; media",
     "Global", "Free", "Public API",
     "https://api.gbif.org/v1",
     "JSON / REST. Darwin Core standard.", "Continuous",
     "High for occurrence data; provenance quality varies by dataset",
     "No", "None", "Low",
     "Search / Compliance", "Governance Scan", "High",
     "GBIF occurrence API is foundational for species-to-country mapping. "
     "Key endpoint: /occurrence/search?scientificName=",
     "[SRS-REQ] Platform species search must query GBIF to auto-identify "
     "provider countries, triggering the correct national ABS pathway."],

    ["GenBank (NCBI)",
     "US National Center for Biotechnology Information's sequence database. "
     "Primary global repository for DNA sequence data (DSI). "
     "Critical for Digital Sequence Information governance context.",
     "DNA/RNA sequences; genome records; protein data; literature links; "
     "BioProject and BioSample metadata",
     "Global", "Free", "Public API (strong)",
     "https://www.ncbi.nlm.nih.gov/genbank",
     "FASTA; GenBank flat file; JSON via Entrez API", "Continuous",
     "Very high — >2 trillion base pairs; authoritative",
     "No", "None", "Low",
     "Search / Due Diligence", "Governance Scan", "High",
     "Critical for DSI discussions: sequences derived from genetic resources "
     "accessed under ABS obligations may be deposited here. Provenance "
     "tracking from physical sample to GenBank record is a key SRS challenge.",
     "[SRS-REQ] Platform must track the GenBank accession number for any "
     "digital sequence information derived from a resource accessed through "
     "the platform, linking physical sample provenance to DSI record."],

    ["ENA (European Nucleotide Archive)",
     "EMBL-EBI's European nucleotide sequence archive. EU-aligned DSI source "
     "and mirror of GenBank within the International Nucleotide Sequence "
     "Database Collaboration (INSDC).",
     "Sequence data; genome assemblies; metagenomics; sample metadata; "
     "ENA study records",
     "Global", "Free", "Public API (strong)",
     "https://www.ebi.ac.uk/ena",
     "Standardised genomic formats; REST API", "Continuous",
     "Very high — EU-preferred genomic archive",
     "No", "None", "Low",
     "Search / Due Diligence", "Governance Scan", "Very High",
     "EU-aligned alternative to GenBank for DSI provenance tracking. "
     "Particularly relevant for EU-funded research outputs and EU regulatory "
     "reporting on DSI utilisation.",
     "[SRS-REQ] For EU-based users and EU-funded research, ENA accession "
     "numbers should be the preferred DSI provenance reference alongside "
     "GenBank."],

    ["DDBJ (DNA Data Bank of Japan)",
     "Japan's nucleotide sequence database; third member of the INSDC "
     "alongside GenBank and ENA. Relevant for Asian provider country DSI.",
     "DNA sequences; genome data; standard genomic formats",
     "Global", "Free", "API available",
     "https://www.ddbj.nig.ac.jp",
     "Standard genomic formats; DDBJ REST API", "Continuous",
     "High — complements GenBank/ENA for Asian sequence data",
     "No", "None", "Low",
     "Search", "Governance Scan", "Medium",
     "Complementary to GenBank and ENA. Particularly relevant for resources "
     "originating from Asian focus countries (India, Malaysia, Indonesia, "
     "Vietnam, Bhutan).",
     ""],

    ["Genesys PGR",
     "Global portal for plant genetic resources aggregating passport and "
     "characterisation data from CGIAR genebanks and national centres. "
     "Links to ITPGRFA SMTA status.",
     "Accession passport data (MCPD); characterisation traits; genebank "
     "source; SMTA availability flag; crop species focus",
     "Global (genebank holdings)", "Free", "Public API",
     "https://www.genesys-pgr.org/en/a/v2",
     "JSON / REST. Multi-Crop Passport Descriptors (MCPD).", "Regular",
     "High for covered genebanks; limited for non-member collections",
     "No", "None", "Low",
     "Search / Compliance", "Governance Scan / Partner Scan", "High",
     "Critical for determining if a target crop accession is covered by "
     "ITPGRFA Multilateral System (SMTA applies instead of national ABS).",
     "[SRS-REQ] For plant genetic resource requests, platform must query "
     "Genesys to check ITPGRFA SMTA coverage; if covered, route to SMTA "
     "workflow; if not, route to national ABS workflow."],

    ["GGBN Data Portal",
     "Global Genome Biodiversity Network portal aggregating tissue and DNA "
     "sample records. Contains the GGBN permit vocabulary — the emerging "
     "data standard for ABS compliance metadata.",
     "Specimen and sample records; permit vocabulary fields (PIC, MAT, MTA); "
     "collection metadata; provenance; post-Nagoya compliance flags",
     "Global (member institutions)", "Free", "Public API",
     "https://data.ggbn.org/ggbn_portal/api",
     "JSON. GGBN Data Standard / ABCDE extension.", "Regular",
     "High for GGBN members; mandatory for post-Nagoya collections",
     "No", "None", "Medium",
     "Compliance / Due Diligence", "Governance Scan", "High",
     "GGBN permit vocabulary (PIC, MAT, MTA) is the authoritative data "
     "standard for ABS compliance metadata across biological collections.",
     "[SRS-REQ] All genetic resource records in the platform must carry "
     "GGBN permit vocabulary metadata as mandatory fields. Platform data "
     "model must align with the GGBN Data Standard."],

    ["IUCN Red List",
     "Authoritative global database on the conservation status of species. "
     "Essential for sustainability screening of bioprospecting targets.",
     "Conservation status (LC/NT/VU/EN/CR); population trend; habitat; "
     "geographic range; threat factors; use and trade records",
     "Global", "Free (API with key)", "Auth API",
     "https://apiv3.iucnredlist.org",
     "JSON / REST", "Annual updates",
     "High — authoritative conservation status",
     "Yes (free API key registration)", "None", "Low",
     "Due Diligence / Search", "Governance Scan", "High",
     "Critical sustainability screening: platform must flag resources with "
     "CR/EN/VU status before matching.",
     "[SRS-REQ] Species screening must query IUCN Red List; resources with "
     "CR or EN status must require additional sustainability justification "
     "before the matching workflow proceeds."],

    ["Catalogue of Life",
     "Authoritative global checklist of all known species (~2M species). "
     "Used for taxonomic name resolution and backbone consistency.",
     "Species names; accepted taxonomy; synonyms; family/genus hierarchy; "
     "authorship",
     "Global", "Free", "Limited API",
     "https://www.catalogueoflife.org",
     "Structured taxonomy; API v1", "Annual",
     "High — editorially curated by 500+ expert groups",
     "No", "None", "Low",
     "Search", "Partner Scan", "High",
     "Taxonomic backbone for name resolution alongside POWO (plants) and "
     "GBIF. Important for handling synonym variants in species search.",
     "[SRS-REQ] Platform species name resolution should cross-reference "
     "Catalogue of Life for all taxonomic groups beyond plants."],

    ["Plants of the World Online (POWO — Kew)",
     "Comprehensive database of the world's flora maintained by Royal Botanic "
     "Gardens, Kew. Authoritative for plant taxonomy and nomenclature.",
     "Accepted plant names; synonyms; distribution maps; morphological "
     "descriptions; image records",
     "Global", "Free", "Public API",
     "https://powo.science.kew.org/api/2",
     "JSON / REST", "Continuous",
     "Very high — editorial standard from Kew",
     "No", "None", "Low",
     "Search", "Partner Scan", "Medium",
     "Preferred plant taxonomy backbone over generic Catalogue of Life "
     "for plant species. Query first for plant resources before GBIF.",
     "[SRS-REQ] Platform species search for plant resources should resolve "
     "names against POWO before querying GBIF, to handle synonyms and "
     "regional name variants."],

    ["Protected Planet (WDPA)",
     "UNEP-WCMC global database of protected areas. 250,000+ sites with "
     "spatial boundaries. Monthly updates.",
     "Protected area boundaries; IUCN category; governance type; country; "
     "area designation; management authority",
     "Global", "Free", "API available",
     "https://www.protectedplanet.net",
     "Spatial datasets (GeoJSON / shapefile); REST API", "Monthly",
     "High — official UN protected areas dataset",
     "No", "None", "Low",
     "Due Diligence / Search", "Governance Scan", "High",
     "Critical for determining whether a target resource originates from "
     "a protected area, which triggers additional permit requirements in "
     "most provider countries.",
     "[SRS-REQ] Species occurrence data from GBIF must be cross-referenced "
     "against WDPA to flag resources originating from protected areas, "
     "triggering additional permit guidance."],

    ["OBIS (Ocean Biodiversity Information System)",
     "IOC-UNESCO global repository of marine species occurrence data. "
     "Relevant for marine genetic resources (MGR) from focus countries "
     "with significant coastal/marine biodiversity.",
     "Marine species occurrence records; sampling event data; collection "
     "metadata; environmental parameters",
     "Global", "Free", "API available",
     "https://obis.org",
     "Standardised biodiversity data; Darwin Core; REST API", "Ongoing",
     "High for OBIS member institutions; growing coverage",
     "No", "None", "Low",
     "Search", "Governance Scan", "Medium",
     "Relevant for Trinidad & Tobago, Indonesia, Malaysia, and other "
     "focus countries with significant marine resources. Marine genetic "
     "resources (MGR) are an increasingly important ABS frontier.",
     "[SRS-REQ] Platform species search must include OBIS as the marine "
     "biodiversity data source for any resources identified as marine-origin."],

    ["FAO AGRIS",
     "FAO's international system for agricultural science and technology "
     "information. Linked data with millions of records on agricultural research.",
     "Agricultural research publications; project data; agricultural "
     "innovation records; linked open data",
     "Global", "Free", "Limited API",
     "https://agris.fao.org",
     "Linked data; Dublin Core metadata", "Ongoing",
     "Moderate — coverage varies by publishing institution",
     "No", "None", "Low",
     "Search", "Partner Scan", "Medium",
     "Useful for identifying research on specific crops and agricultural "
     "resources in focus countries. Agri-bioinnovation relevance.",
     ""],

    ["World Flora Online",
     "Consortium-maintained database of all global plant species (~350k). "
     "Complementary to POWO for comprehensive plant coverage.",
     "Global plant species names; distribution; morphological descriptions",
     "Global", "Free", "Limited API",
     "http://www.worldfloraonline.org",
     "Structured taxonomy", "Ongoing",
     "High — broad coverage; editorial consortium",
     "No", "None", "Low",
     "Search", "Partner Scan", "Medium",
     "Use as a complement to POWO for plant resources not covered by Kew's "
     "editorial scope.", ""],

    ["BOLD Systems (Barcode of Life)",
     "Global database of DNA barcode reference sequences for species "
     "identification. 500,000+ species barcoded. Useful for species "
     "identity verification.",
     "DNA barcode sequences; specimen records; collection locality; "
     "collector information; institution codes",
     "Global", "Free", "Public API",
     "https://v3.boldsystems.org/index.php/API_Public",
     "JSON / XML", "Continuous",
     "High for barcoded specimens; growing coverage",
     "No", "None", "Low",
     "Search / Due Diligence", "Partner Scan", "Medium",
     "DNA barcoding as proof of species identity useful for provenance "
     "verification. Also identifies which institutions hold reference "
     "specimens.",
     ""],
]

# --- II-D: Southern Institution DBs ---
SOUTH_ROWS = [
    ["Proposed: GENE-LINK Verified Partner Registry",
     "Custom database to be built. Verified profiles of Southern research "
     "institutions, genebanks, and bio-innovation organisations that have "
     "completed GENE-LINK onboarding, identity verification, and ABS "
     "compliance assessment.",
     "Institution profiles; resource specialisations; ABS compliance capacity; "
     "ORCID-linked researcher profiles; verification status; review history",
     "Global South (15 focus countries + expansion)",
     "Platform-internal (private)", "Platform database",
     "Proposed — to build", "Custom schema",
     "Real-time", "High (platform-owned, verified data)",
     "N/A", "N/A", "High (to build)",
     "Matchmaking", "Partner Scan", "Very High",
     "This is the core provider-side asset. Must be designed before "
     "platform development begins.",
     "[SRS-REQ] Platform must maintain a verified partner registry with "
     "mandatory fields: institution type, country, resource specialisation, "
     "ABS compliance status, and at least one ORCID-linked researcher "
     "contact."],

    ["CGIAR Genebank Platform",
     "Network of 15 CGIAR genebanks holding 700,000+ crop and forage "
     "accessions. Exposed via Genesys PGR. Most resources governed by "
     "ITPGRFA SMTA.",
     "Accession data (via Genesys); centre mandates; focus crops; distribution "
     "records; quarantine status",
     "Global South: Ethiopia, Philippines, Colombia, Nigeria, Mexico, Kenya, "
     "India, Peru, Syria",
     "Free (via Genesys API)", "Public API (via Genesys)",
     "https://www.genesys-pgr.org / https://www.cgiar.org/initiative/genebanks",
     "MCPD / DOI-linked accessions", "Regular",
     "High — FAO QMS compliance; DOI-assigned accessions",
     "No", "None", "Low",
     "Matchmaking / Search", "Partner Scan / Governance Scan", "High",
     "Largest single source of agricultural genetic resources in the Global "
     "South. Key centres: ILRI (Ethiopia), IRRI (Philippines), IITA (Nigeria), "
     "CIMMYT (Mexico), CIP (Peru). ITPGRFA SMTA applies to most crops.",
     "[SRS-REQ] CGIAR centres must be pre-populated as verified provider "
     "institution profiles in the platform for relevant focus countries."],

    ["CIRAD Platforms in Partnership (dPs)",
     "CIRAD's network of joint research and training platforms in Africa, "
     "Asia, and Latin America. Interface between European scientists and "
     "regional research laboratories.",
     "Platform descriptions; regional contacts; focus topics; partner "
     "institution lists; ongoing research themes",
     "Africa / Asia / Latin America (CIRAD presence countries)",
     "Free (directory)", "Portal / Manual",
     "https://www.cirad.fr/en/worldwide/platforms-in-partnership",
     "HTML / unstructured", "As updated",
     "Good for coverage; varies by platform",
     "No", "None", "Medium",
     "Matchmaking", "Partner Scan", "Medium",
     "Relevant dPs: ASEA/GREASE (SE Asia), AGROFORESTA (Latin America). "
     "No API — manual extraction required. CIFOR-ICRAF partnership should "
     "be explored.",
     "[SRS-REQ] CIRAD dP partner institutions should be candidates for "
     "Southern provider profiles; CIFOR-ICRAF partnership should be "
     "explored as an onboarding channel."],

    ["BioInnovation Africa (BIA) — abs-biotrade.info",
     "GIZ-managed repository of African biodiversity-based value chains, "
     "ABS best practices, and EU-Africa partnership case studies. "
     "7,400+ field operators documented.",
     "Value chain case studies; ABS manuals; MAT negotiation training; "
     "partner institution contacts; benefit-sharing best practices",
     "Sub-Saharan Africa (primary)",
     "Free", "Portal / Manual",
     "https://www.abs-biotrade.info",
     "HTML / PDF", "As updated (project-driven)",
     "Practitioner-quality; limited machine-readable extraction",
     "No", "None", "Medium",
     "Matchmaking / Capacity Dev", "Governance Scan", "High",
     "Best available single source for African ABS-compliant value chain "
     "documentation. GIZ data-sharing agreement should be pursued.",
     "[SRS-REQ] Platform guidance module for African provider countries "
     "should integrate BIA ABS manuals as authoritative practitioner "
     "resources via formal GIZ content partnership."],

    ["ORCID (Researcher Identity)",
     "Persistent researcher identifier infrastructure linked to publications, "
     "affiliations, and funding. De facto global researcher identity standard.",
     "Researcher IDs; publication DOIs; employment history; funding grants; "
     "peer review records",
     "Global", "Free", "Public API (read)",
     "https://pub.orcid.org/v3.0",
     "JSON / XML", "Continuous (user-maintained)",
     "High for researchers who maintain their ORCID profile",
     "No (read) / Yes (write — member API)",
     "None (read)", "Low",
     "Matchmaking", "Partner Scan", "High",
     "Infrastructure for researcher identity in the platform. Southern "
     "institution researchers should be encouraged to register ORCID IDs "
     "as part of onboarding.",
     "[SRS-REQ] Platform researcher profiles must support ORCID "
     "integration for identity verification and publication record linkage."],

    ["Indigenous Peoples Atlas (TK Data Gap Reference)",
     "Example knowledge platform illustrating the gap in machine-readable "
     "Traditional Knowledge (TK) databases for indigenous providers. "
     "Included as a gap-marker, not an integration target.",
     "Indigenous knowledge documentation; cultural heritage data (example "
     "only — limited geographic scope)",
     "Regional (Canada example)", "Open", "Portal (no API)",
     "https://indigenouspeoplesatlasofcanada.ca",
     "Mixed; unstructured", "Static",
     "Limited; example only",
     "No", "None", "N/A",
     "Reference (gap illustration)", "Partner Scan", "Low",
     "Illustrates the absence of any global, standardised, machine-readable "
     "TK database. GENE-LINK must develop its own TK documentation approach "
     "in consultation with indigenous communities.",
     "[SRS-REQ] Platform must develop a bespoke TK documentation module — "
     "no existing database can be integrated. This module must be designed "
     "in consultation with indigenous community representatives."],
]


# ---------------------------------------------------------------------------
# Sheet II-E — Content & Knowledge Assets (merged E-Learning + Knowledge)
# ---------------------------------------------------------------------------

CONTENT_COLS = [
    "Content Name / Title",
    "Content Type\n(Course / Module / Toolkit / Template\n"
    "/ Standard / Framework / Case Study / Proposed)",
    "URL / Access Link",
    "Provider Organisation",
    "Format",
    "Description",
    "ABS Coverage\n(Y / Partial / N)",
    "IP for Bio Resources\n(Y / Partial / N)",
    "Commercialisation\n(Y / Partial / N)",
    "North-South Partnerships\n(Y / Partial / N)",
    "Responsible Sourcing\n(Y / Partial / N)",
    "Target Audience",
    "Language(s)",
    "Cost",
    "Duration / Effort",
    "Geographic Focus",
    "Year / Version",
    "Quality / Credibility",
    "Usability\n(a=direct / b=adapt / c=develop)",
    "Integration Method\n(Link / Embed / Download / Hosted)",
    "Licence / Permission",
    "Priority for Integration",
    "Dedup Note\n(if covered elsewhere in workbook)",
    "SRS Requirement",
]

CONTENT_ROWS = {
    "E-LEARNING COURSES & TRAINING MODULES": [
        ["ABS Clearing-House Training Materials",
         "Online Resources / Guidance",
         "https://absch.cbd.int",
         "CBD Secretariat",
         "Online resources; guidance documents; training modules",
         "Authoritative ABS compliance guidance and training materials "
         "produced by the CBD Secretariat. Covers permits, legal frameworks, "
         "and IRCC processes.",
         "Y", "N", "Partial", "Y", "Y",
         "Governments; regulators; researchers",
         "EN + multilingual", "Free", "Self-paced", "Global", "2023",
         "High (official CBD)", "a",
         "Link at compliance workflow steps",
         "Open", "High",
         "ABSCH as a DATA SOURCE is captured in II-A. This entry covers "
         "only the training/guidance content dimension.",
         "[SRS-REQ] Platform compliance module must link to ABSCH training "
         "materials at each relevant ABS workflow step."],

        ["ABS Capacity Development Initiative (GIZ)",
         "Toolkit / Training Module",
         "https://www.abs-initiative.info",
         "GIZ / International partners",
         "Toolkits; training modules; practitioner guides",
         "Practical ABS capacity development materials covering access, "
         "benefit-sharing, partnership structures, and negotiation. "
         "Highly relevant and practitioner-tested.",
         "Y", "Partial", "Y", "Y", "Y",
         "Practitioners; policymakers; researchers",
         "EN", "Free", "Short modules", "Global South", "2022–2024",
         "High", "a",
         "Embed in guidance module for focus countries",
         "Open (confirm with GIZ)", "High",
         "Note: BIA materials (abs-biotrade.info) are the Africa-specific "
         "branch of this initiative. BIA captured separately below.",
         "[SRS-REQ] Platform guidance module must integrate GIZ ABS CDI "
         "materials as the core practitioner toolkit for non-African contexts."],

        ["WIPO eLearning: IP and Genetic Resources",
         "Online Course",
         "https://welc.wipo.int",
         "WIPO",
         "Self-paced online course with assessments",
         "Comprehensive 20-30 hour course covering intellectual property "
         "law for genetic resources and traditional knowledge. Directly "
         "usable for EU company legal teams and Southern institution staff.",
         "Partial", "Y", "N", "Partial", "Partial",
         "Lawyers; policymakers; researchers",
         "EN + others", "Free", "20–30 hrs", "Global", "2023",
         "High", "a",
         "Link at IPR/due diligence workflow steps; embed certificate in "
         "user profile as a trust signal",
         "Open (WIPO licence)", "High",
         "",
         "[SRS-REQ] Completion of the WIPO IP and Genetic Resources course "
         "should be offered as a verified credential on user profiles, "
         "serving as a trust signal for IP awareness."],

        ["UNDP BioTrade Training Modules",
         "Training Module",
         "https://www.biotrade.org",
         "UNCTAD / UNDP",
         "Training modules; practitioner guides",
         "Training covering BioTrade principles, sustainable value chains, "
         "and biodiversity-based business models. Strong relevance to "
         "GENE-LINK's bioeconomy framing.",
         "Partial", "Y", "Partial", "Y", "Y",
         "SMEs; policymakers; practitioners",
         "EN/ES", "Free", "Medium", "Global South", "2022",
         "High", "a",
         "Embed in capacity building module for Southern providers",
         "Open", "High",
         "UNCTAD BioTrade e-learning (captured below) is a related but "
         "distinct resource from the same organisation.",
         ""],

        ["UNCTAD BioTrade E-Learning",
         "Online Course / Training",
         "https://unctad.org",
         "UNCTAD",
         "Online training; short modules",
         "E-learning focused on BioTrade principles, biodiversity value "
         "chains, and sustainable sourcing from UNCTAD.",
         "Partial", "Y", "Partial", "Y", "Y",
         "SMEs; policymakers", "EN/ES", "Free", "Medium", "Global South",
         "2022", "High", "a",
         "Link from platform's BioTrade-specific guidance sections",
         "Open", "High",
         "Related to UNDP BioTrade modules above (different org, same "
         "BioTrade initiative). Both should be linked, not duplicated.",
         ""],

        ["ABS Contracting & Benefit-Sharing Toolkits",
         "Toolkit / Operational Tool",
         "Various (ABS Initiative; NGOs; national governments)",
         "ABS Capacity Development Initiative / NGOs",
         "PDF toolkits; model contract collections",
         "Collection of practical ABS contracting toolkits, model clauses, "
         "and benefit-sharing calculation tools from various ABS practitioners "
         "and the GIZ ABS Initiative.",
         "Y", "Y", "Y", "Y", "Y",
         "Practitioners; legal teams",
         "EN", "Free", "Short", "Global South", "2022–2024",
         "High", "a",
         "Embed as downloadable templates in MTA/agreement workflow",
         "Open (varies — confirm per toolkit)", "Very High",
         "",
         "[SRS-REQ] Platform MTA workflow must include links to ABS "
         "contracting toolkits as starting-point templates before users "
         "engage with OpenMTA or custom GENE-LINK templates."],

        ["Nagoya Protocol Hub — Country Guides",
         "Guidance / Training",
         "https://nagoyaprotocol-hub.de/toolsandresources-abs",
         "German Nagoya Protocol Hub (government-funded)",
         "PDF country guides; web articles",
         "Country-specific ABS compliance guides (Easy ABS Guides) and "
         "practitioner training content. The most user-friendly country-level "
         "compliance guidance available.",
         "Y", "N", "N", "Partial", "Y",
         "Researchers; companies; practitioners",
         "EN/DE/ES (varies)", "Free", "Short (per-country)", "Global",
         "2023–2025", "High", "a",
         "Link/embed country guides at provider-country selection step",
         "Open (attribution)", "High",
         "Nagoya Protocol Hub as a DATA SOURCE is captured in II-A. "
         "This entry covers the country guide training content.",
         "[SRS-REQ] When a user selects a provider country, platform must "
         "surface the relevant Nagoya Protocol Hub country guide alongside "
         "live ABSCH data."],

        ["BioInnovation Africa — Training Packages",
         "Training Module / Toolkit",
         "https://www.abs-biotrade.info",
         "GIZ / BioInnovation Africa",
         "PDF manuals; workshop materials; case study documents",
         "GIZ-produced training materials on ABS compliance, MAT "
         "negotiation, and African biodiversity-based value chains for "
         "African provider country contexts.",
         "Y", "Partial", "Y", "Y", "Y",
         "African bio-innovators; EU sourcing teams; practitioners",
         "EN/FR (varies)", "Free", "Medium", "Africa", "2022–2024",
         "High", "a",
         "Embed in platform guidance module for African provider countries",
         "Open (GIZ data-sharing agreement recommended)", "High",
         "BIA as a SOUTHERN INSTITUTION DATABASE is captured in II-D. "
         "This entry covers only the training content dimension.",
         "[SRS-REQ] Platform guidance for African provider countries must "
         "integrate BIA training packages via formal GIZ content partnership."],

        ["FAO E-Learning: ITPGRFA / Plant Genetic Resources",
         "Online Course",
         "https://elearning.fao.org",
         "FAO",
         "Online modules with assessments",
         "FAO training modules on the ITPGRFA treaty obligations, the "
         "Standard Material Transfer Agreement (SMTA), and sustainable "
         "use of plant genetic resources.",
         "Partial", "Partial", "Partial", "Y", "Y",
         "Researchers; genebank managers; policy officials",
         "EN/FR/ES", "Free", "5–10 hrs", "Global", "2022",
         "High", "a",
         "Link from platform SMTA workflow step",
         "Open", "Medium",
         "FAO AGRIS as a DATA SOURCE is in II-C. This entry covers the "
         "ITPGRFA learning content only.",
         "[SRS-REQ] When the platform SMTA workflow is triggered, the FAO "
         "ITPGRFA e-learning module must be surfaced for users unfamiliar "
         "with the treaty."],

        ["WIPO GREEN Learning Resources",
         "Case Studies / Resources",
         "https://www.wipo.int/wipogreen",
         "WIPO",
         "Case studies; web resources",
         "Case studies and learning resources on green technology transfer "
         "and North-South partnerships from WIPO GREEN. Adaptable for "
         "GENE-LINK's bio-innovation partnership context.",
         "N", "Partial", "Y", "Y", "Y",
         "Companies; innovators",
         "EN", "Free", "Short", "Global", "2023",
         "High", "b",
         "Adapt for bio-innovation; link from partnership case studies section",
         "Open (WIPO)", "Medium",
         "WIPO GREEN as a PLATFORM is captured in I-A. This entry covers "
         "only the learning/case study content.",
         ""],

        ["OECD RBC Due Diligence Guidance",
         "Guidance / Training",
         "https://mneguidelines.oecd.org",
         "OECD",
         "Policy guidance + training modules",
         "OECD Responsible Business Conduct due diligence guidance and "
         "associated training content. Directly applicable to EU company "
         "sourcing due diligence obligations.",
         "N", "N", "Partial", "Y", "Y",
         "Companies; policymakers",
         "EN", "Free", "Short", "Global", "2023",
         "High", "a",
         "Link from platform due diligence workflow for EU company users",
         "Open", "High",
         "See also: OECD Due Diligence Guidance as a KNOWLEDGE ASSET "
         "in the section below.",
         "[SRS-REQ] Platform due diligence workflow for EU companies must "
         "reference OECD RBC guidance as the responsible sourcing standard."],

        ["UEBT E-Learning (Ethical BioTrade)",
         "Online Course / Corporate Training",
         "https://www.ethicalbiotrade.org/training",
         "UEBT",
         "Online modules; corporate training packages",
         "UEBT training for EU corporate procurement on ethical biodiversity "
         "sourcing, ABS obligations, and supply chain due diligence.",
         "Partial", "N", "Y", "Partial", "Y",
         "EU corporate procurement; sourcing managers",
         "EN/ES", "Paid (membership)", "Medium", "Global", "2023",
         "High", "b",
         "Partner with UEBT; link training for company users",
         "Requires UEBT partnership", "Medium",
         "UEBT as a DATA SOURCE (company directory) is in II-B. "
         "This entry covers only the training content.",
         "[SRS-REQ] GENE-LINK company onboarding should reference UEBT "
         "training and offer UEBT certification as a recommended pathway."],

        ["Coursera: IP for Business",
         "Online Course",
         "https://www.coursera.org",
         "Various universities (via Coursera)",
         "Online course with assessments; certificate",
         "Strong general IP and commercialisation content. Adaptable to "
         "bio-resource context.",
         "N", "Y", "Y", "N", "N",
         "Entrepreneurs; businesses",
         "EN", "Freemium", "20–40 hrs", "Global", "2023",
         "High", "b",
         "Link as supplementary IP learning resource; adapt to bio context",
         "Freemium (Coursera ToS)", "Low",
         "Generic content requiring contextualisation for ABS/bio-innovation.",
         ""],

        ["Custom: GENE-LINK Partner Readiness Course",
         "Proposed / To Develop",
         "Proposed module",
         "GENE-LINK (to develop)",
         "Integrated online course with assessments and certification",
         "No integrated course currently exists combining ABS compliance, "
         "IP for genetic resources, North-South partnership frameworks, and "
         "commercialisation pathways. This is a critical content gap.",
         "Y", "Y", "Y", "Y", "Y",
         "EU companies + Southern research institutions",
         "EN + target languages", "Free (for platform users)",
         "8–12 hrs", "Global", "To develop",
         "N/A (to build)", "c",
         "Hosted on GENE-LINK platform (SCORM-compatible LMS)",
         "GENE-LINK IP", "Very High",
         "",
         "[SRS-REQ] GENE-LINK must develop an integrated Partner Readiness "
         "Course combining ABS, IP, partnerships, and commercialisation — "
         "this is a key platform differentiator that no existing course "
         "provides."],

        ["Custom: ABS + IP + Commercialisation Toolkit",
         "Proposed / To Develop",
         "Proposed toolkit",
         "GENE-LINK (to develop)",
         "Modular downloadable and in-platform toolkit",
         "Fragmented content exists (ABS toolkits, WIPO IP guidelines, UEBT "
         "standards) but no integrated toolkit combining all elements "
         "in a GENE-LINK-specific workflow context.",
         "Y", "Y", "Y", "Y", "Y",
         "SMEs; researchers; legal teams",
         "EN + target languages", "Free (for platform users)",
         "Modular — as needed", "Global", "To develop",
         "N/A (to build)", "c",
         "Hosted on GENE-LINK platform; integrated into compliance workflow",
         "GENE-LINK IP", "Very High",
         "",
         "[SRS-REQ] GENE-LINK must develop a modular integrated toolkit "
         "combining ABS, IP, partnership, and commercialisation tools — "
         "this is the operational companion to the Partner Readiness Course."],
    ],

    "KNOWLEDGE & LEGAL ASSETS": [
        ["ABS Contractual Clauses & Model Agreements (GIZ)",
         "Template / Model Contract",
         "https://www.abs-initiative.info/knowledge-centre",
         "ABS Capacity Development Initiative (GIZ)",
         "Model contracts; downloadable templates; legal clauses",
         "The most comprehensive publicly available collection of ABS model "
         "contract clauses, benefit-sharing agreement templates, and MAT "
         "negotiation guides. Core operational tool for platform MTA module.",
         "Y", "Y", "Y", "Y", "Y",
         "Legal teams; practitioners; researchers",
         "EN", "Free", "Reference",
         "Global", "2022–2024", "Very High",
         "a",
         "Embed as downloadable templates in platform MTA workflow",
         "Likely open — confirm with GIZ", "Very High",
         "",
         "[SRS-REQ] Platform MTA generation module must embed GIZ model "
         "ABS clauses as the default starting point for benefit-sharing "
         "agreement drafting."],

        ["UEBT Responsible Sourcing Standard",
         "Standard / Certification Framework",
         "https://www.ethicalbiotrade.org",
         "UEBT",
         "Standard document; certification framework",
         "UEBT ethical sourcing standard covering biodiversity conservation, "
         "community rights, and fair benefit-sharing. "
         "Key trust signal for EU cosmetics and food sector.",
         "Partial", "N", "Y", "Y", "Y",
         "EU companies; ingredient suppliers",
         "EN", "Free (standard) / Paid (certification)",
         "Reference", "Global", "2020+",
         "Very High", "a",
         "Integrate as a trust indicator in company profiles; link as "
         "a certification pathway",
         "Permission likely required (contact UEBT)", "Very High",
         "UEBT as a DATA SOURCE (company directory) is in II-B. "
         "This entry covers the standard document as a knowledge asset.",
         "[SRS-REQ] Platform must recognise UEBT certification as a verified "
         "trust indicator; the UEBT standard should be embedded as a "
         "responsible sourcing reference in the company onboarding flow."],

        ["OECD Due Diligence Guidance for RBC",
         "Policy / Guidance Document",
         "https://mneguidelines.oecd.org",
         "OECD",
         "Policy guidance document (PDF)",
         "OECD Responsible Business Conduct Due Diligence guidance for "
         "cross-sector responsible sourcing. Internationally recognised "
         "standard for corporate supply chain due diligence.",
         "N", "N", "Partial", "Y", "Y",
         "Companies; policymakers",
         "EN + others", "Free", "Reference", "Global", "2018+",
         "Very High", "a",
         "Embed as reference document in platform due diligence module",
         "Open", "High",
         "See also: OECD RBC e-learning in E-Learning section above.",
         ""],

        ["UN BioTrade Principles and Criteria",
         "Framework / Principles",
         "https://unctad.org/topic/trade-and-environment/biotrade",
         "UNCTAD",
         "Framework document (PDF)",
         "UNCTAD's 7 BioTrade principles and criteria for sustainable, "
         "equitable biodiversity-based trade. Strong alignment with "
         "GENE-LINK's values framework.",
         "Partial", "Y", "Y", "Y", "Y",
         "Companies; policymakers; practitioners",
         "EN/ES", "Free", "Reference", "Global", "2020+",
         "Very High", "a",
         "Embed in platform values/trust layer and company onboarding",
         "Open", "Very High",
         "",
         "[SRS-REQ] GENE-LINK platform values framework must explicitly "
         "reference and embed UNCTAD BioTrade Principles as the responsible "
         "sourcing standard alongside UEBT."],

        ["WIPO IP Guidelines for Genetic Resources",
         "Guidelines / Reference",
         "https://www.wipo.int",
         "WIPO",
         "PDF guidelines",
         "WIPO's intellectual property guidelines specific to genetic "
         "resources, traditional knowledge, and folklore. Critical reference "
         "for the platform's IP guidance module.",
         "Partial", "Y", "Partial", "Partial", "Partial",
         "Legal teams; policymakers; researchers",
         "EN", "Free", "Reference", "Global", "2022+",
         "Very High", "a",
         "Embed in platform IP guidance module; link from due diligence "
         "workflow",
         "Open (WIPO)", "Very High",
         "See also: WIPO eLearning course in E-Learning section above.",
         "[SRS-REQ] Platform IP guidance module must embed WIPO IP "
         "Guidelines for Genetic Resources as the authoritative reference."],

        ["GIZ Value Links Methodology",
         "Methodology / Toolkit",
         "https://www.giz.de",
         "GIZ",
         "Methodology guide; workshop tools",
         "GIZ's Value Links methodology for value chain analysis and "
         "partnership structuring. Adaptable for analysing bio-innovation "
         "value chains in focus countries.",
         "N", "N", "Y", "Y", "Y",
         "Development practitioners; programme designers",
         "EN", "Free (contact GIZ)", "Reference", "Global", "Updated",
         "High", "b",
         "Adapt into GENE-LINK value chain mapping tools for partner onboarding",
         "Permission required (contact GIZ)", "High",
         "",
         ""],

        ["CIFOR-ICRAF Agroforestry Case Studies",
         "Case Studies",
         "https://www.worldagroforestry.org",
         "CIFOR-ICRAF",
         "Published case studies and reports",
         "Agroforestry and value chain case studies from CIFOR-ICRAF. "
         "Relevant as GENE-LINK's institutional home and for agroforestry-"
         "based resource value chains.",
         "N", "N", "Y", "Y", "Y",
         "Researchers; practitioners",
         "EN", "Open", "Reference", "Global South", "Ongoing",
         "High", "a",
         "Link from platform's bio-innovation case study library",
         "Open", "High",
         "CIFOR-ICRAF is GENE-LINK's institutional host — internal content "
         "partnership should be formalised.",
         "[SRS-REQ] Platform case study library must include CIFOR-ICRAF "
         "agroforestry case studies as foundational content given the "
         "institutional relationship."],

        ["CBD Voluntary Guidelines on ABS Compliance",
         "Guidelines / Reference",
         "https://www.cbd.int",
         "CBD Secretariat",
         "PDF guidelines",
         "CBD Voluntary Guidelines on ABS compliance for private sector "
         "and research institutions. Complementary to the Nagoya Protocol "
         "text and the ABSCH regulatory framework.",
         "Y", "N", "Partial", "Y", "Y",
         "Companies; researchers; practitioners",
         "EN + multilingual", "Free", "Reference", "Global", "2018+",
         "High", "a",
         "Link from platform compliance overview and due diligence module",
         "Open", "High",
         "ABSCH as a DATA SOURCE is in II-A. Nagoya Protocol Hub country "
         "guides (E-Learning above) are the operational derivative. This "
         "entry covers the voluntary guidelines document only.",
         ""],

        ["Nagoya Protocol Text and Explanatory Guides",
         "Legal Framework / Reference",
         "https://www.cbd.int/abs/instruments",
         "CBD Secretariat",
         "Legal treaty text; PDF explanatory guides",
         "The Nagoya Protocol treaty text and CBD explanatory guides. "
         "Foundational legal reference for all ABS compliance content.",
         "Y", "N", "N", "Y", "N",
         "Legal teams; policymakers; compliance officers",
         "EN + multilingual", "Free", "Reference", "Global", "2014+",
         "High", "a",
         "Link from platform's legal reference library",
         "Open", "High",
         "ABSCH as a DATA SOURCE (containing regulatory measures derived "
         "from the Protocol) is in II-A. This entry covers only the "
         "treaty text as a knowledge reference asset.",
         ""],

        ["Custom: GENE-LINK Partnership Agreement Template",
         "Proposed / To Develop",
         "Proposed template",
         "GENE-LINK (to develop)",
         "Standardised contract template; in-platform document generation",
         "No integrated partnership agreement template currently exists "
         "combining ABS, IP, benefit-sharing, and commercialisation clauses "
         "in a single GENE-LINK-context document. Critical platform gap.",
         "Y", "Y", "Y", "Y", "Y",
         "EU companies + Southern research institutions",
         "EN + target languages", "Free (for platform users)",
         "Template (30–60 min to complete)", "Global", "To develop",
         "N/A (to build)", "c",
         "Hosted on platform; generated via MTA module; e-signature enabled",
         "GENE-LINK IP", "Very High",
         "",
         "[SRS-REQ] Platform must develop a GENE-LINK-specific partnership "
         "agreement template combining ABS, IP, benefit-sharing, and "
         "commercialisation clauses — this is a core platform asset and "
         "key differentiator."],

        ["Custom: GENE-LINK Deal Pathway Guide",
         "Proposed / To Develop",
         "Proposed guide",
         "GENE-LINK (to develop)",
         "Step-by-step interactive guide; UX-integrated",
         "A guided, step-by-step partnership pathway from first match to "
         "executed agreement — integrating ABS compliance, MTA generation, "
         "due diligence steps, and benefit-sharing setup in one flow. "
         "Directly bridges the platform UX and the learning content.",
         "Y", "Y", "Y", "Y", "Y",
         "EU companies + Southern research institutions (all users)",
         "EN + target languages", "Free (for platform users)",
         "Interactive (varies by user path)", "Global", "To develop",
         "N/A (to build)", "c",
         "Fully integrated into platform UX; not a standalone document",
         "GENE-LINK IP", "Very High",
         "",
         "[SRS-REQ] Platform UX must embed the Deal Pathway Guide as "
         "an interactive guided workflow — not a downloadable PDF — "
         "bridging the compliance, MTA, and due diligence steps."],

        ["UNDP BIOFIN Materials",
         "Reports / Tools",
         "https://www.biofin.org",
         "UNDP",
         "Reports; methodology tools",
         "UNDP Biodiversity Finance Initiative materials on financing "
         "biodiversity conservation. Contextual background for GENE-LINK's "
         "benefit-sharing and investment framing.",
         "N", "N", "Y", "Y", "Partial",
         "Policymakers; finance practitioners",
         "EN", "Free", "Reference", "Global", "Ongoing",
         "High", "b",
         "Link from platform's contextual resources / background section",
         "Open", "Low",
         "", ""],

        ["Agrinatura Knowledge Outputs",
         "Reports / Case Studies",
         "Contact Agrinatura",
         "Agrinatura",
         "Reports; case studies; partnership documentation",
         "Knowledge outputs from Agrinatura — the European consortium of "
         "agricultural research organisations for international development "
         "that includes GENE-LINK consortium members.",
         "N", "N", "Y", "Y", "Partial",
         "Researchers; development practitioners",
         "EN", "Varies", "Reference", "Global", "Ongoing",
         "Medium–High", "b",
         "Link from platform's partner knowledge section",
         "Varies — internal partner", "Medium",
         "Agrinatura is an internal consortium partner — content access "
         "should be facilitated through the project partnership.",
         ""],
    ],
}


# ---------------------------------------------------------------------------
# Sheet II-F — In-Country Support (empty template from Dominique)
# ---------------------------------------------------------------------------

INCOUNTRY_COLS = [
    "Country", "Organisation Name", "URL / Source", "Organisation Type",
    "Sector Focus", "Services Offered (summary)",
    "Legal / IP Services?", "Commercialisation Services?",
    "North-South Partnership Support?", "ABS / Biodiversity Expertise?",
    "Target Clients", "Geographic Reach", "Scale / Size",
    "Funding / Sustainability Model", "Potential Collaboration Role",
    "Gap vs. Southern Institution Needs",
    "Key Contact / Informant", "Priority for KII",
    "Added By", "Date Added", "Notes",
]


# ---------------------------------------------------------------------------
# Sheet III — Functionalities & Requirements
# ---------------------------------------------------------------------------

FUNC_COLS = [
    "Functionality Area",
    "Functionality Name",
    "Description",
    "User Roles\n(EU User / Southern Provider / Admin / All)",
    "Public or Private\n(Public / Authenticated / Admin-Only)",
    "Supporting Data Sources\n(from Part II)",
    "Platform Analogues\n(from Part I)",
    "Implementation Complexity\n(Low / Med / High)",
    "Privacy & Security Notes",
    "Due Diligence Dimension\n(Y/N — describe)",
    "Priority\n(High / Med / Low)",
    "SRS Requirement",
]

FUNC_ROWS = {
    "ACCESS & DISCOVERY": [
        ["Access & Discovery", "Species & Resource Search",
         "Users search for a genetic resource or species by scientific name, "
         "common name, or resource type. Platform resolves taxonomic synonyms "
         "(POWO, Catalogue of Life), returns provider countries (GBIF), "
         "ITPGRFA SMTA coverage (Genesys), CITES listing, IUCN conservation "
         "status, and applicable ABS regime.",
         "EU User / Southern Provider",
         "Public (basic); Authenticated (compliance detail)",
         "GBIF, Genesys PGR, POWO, Catalogue of Life, BOLD, IUCN Red List, "
         "CITES, Protected Planet (WDPA)",
         "ABSCH (species/country lookup)",
         "Medium",
         "Do not expose sensitive location data for CR/EN species. "
         "Follow GBIF coordinate obscuring policy.",
         "Y — CITES and IUCN screening as part of search",
         "High",
         "[SRS-REQ] Species search must integrate GBIF (provider country), "
         "Genesys (SMTA), CITES (permit flag), IUCN (conservation status), "
         "POWO (name resolution), and WDPA (protected area flag) in one "
         "unified results view."],

        ["Access & Discovery", "Provider Institution Directory",
         "Searchable directory of Southern research institutions and genebanks. "
         "Profiles include: institution type, resource focus, country, ABS "
         "compliance capacity, ORCID-linked researcher contacts, and "
         "verified status.",
         "EU User / All",
         "Public (profiles); Authenticated (contact details)",
         "CGIAR Genebank, CIRAD dPs, BioInnovation Africa, ORCID, "
         "CGHSS AbsMaster, Custom Partner Registry",
         "WIPO MATCH (offer/seek model), ABS Connect, LinkedIn (profile model)",
         "Medium",
         "Contact details of institution staff must be authenticated-access "
         "only to prevent spam.",
         "N",
         "High",
         "[SRS-REQ] Provider directory must support filtering by country, "
         "resource type, sector expertise, and ABS compliance status."],

        ["Access & Discovery", "EU Company Directory (User Profiles)",
         "Searchable directory of EU private sector companies active in "
         "cosmetics, pharma, food, biotech, and specialty ingredients. "
         "Profiles include sector, R&D interests, UEBT certification, and "
         "ABS compliance history.",
         "Southern Provider / Admin",
         "Authenticated (basic profiles public; details authenticated)",
         "John Snow Labs Biotech DB, UEBT Directory, Cosmetics Europe, "
         "CORDIS, ORCID, OpenCorporates",
         "LinkedIn (progressive profile), b2match (structured fields), "
         "Airbnb (trust signals)",
         "High",
         "GDPR compliance review required for all profile data. Companies "
         "must control profile visibility.",
         "Y — UEBT certification and ABS compliance history as trust signals",
         "High",
         "[SRS-REQ] EU company profiles must support UEBT certification, "
         "sectoral tags, and ABS compliance history. GDPR review required. "
         "OpenCorporates verification must be a prerequisite for verified status."],

        ["Access & Discovery", "Request-Driven Resource Matching",
         "EU company posts a resource need (species type, bioactive target, "
         "sector, quantity, geographic preference). Platform matches the need "
         "against provider institution profiles and resource offers. "
         "Follows CTCN's request-response model with added ABS routing.",
         "EU User",
         "Authenticated",
         "Custom Opportunity Pipeline, Custom Partner Registry, GBIF",
         "CTCN (request-response model), WIPO GREEN (need posting)",
         "High",
         "Resource need postings must not reveal proprietary R&D intent to "
         "non-matched parties.",
         "N",
         "High",
         "[SRS-REQ] Platform must support demand-driven matching: EU company "
         "posts a resource need; platform matches against Southern provider "
         "offers using ABS-aware filters and relevance scoring."],
    ],

    "VERIFICATION & DUE DILIGENCE": [
        ["Verification & Due Diligence", "ABS Pre-Flight Compliance Check",
         "When a user identifies a target resource and provider country, "
         "platform runs an automated pre-flight check: queries ABSCH for "
         "NFP/CNA details, IRCC issuance status, applicable permits, and "
         "ITPGRFA SMTA coverage. Surfaces the correct compliance pathway "
         "and generates a time-stamped compliance record.",
         "EU User",
         "Authenticated",
         "ABSCH API, Genesys PGR, SisGen guidance, NBA e-Filing, "
         "CGHSS AbsMaster, Nagoya Protocol Hub guides",
         "ABSCH, Nagoya Protocol Hub (walkthrough model)",
         "High",
         "Compliance records must be time-stamped and immutable for audit. "
         "Must include a clear disclaimer that users remain responsible "
         "for their own due diligence.",
         "Y — core due diligence function",
         "High",
         "[SRS-REQ] Pre-flight compliance check must produce a structured, "
         "time-stamped compliance record per transaction covering: provider "
         "country, resource, applicable regime, NFP/CNA details, and "
         "required next steps."],

        ["Verification & Due Diligence", "IRCC Verification",
         "Automated verification that a genetic resource has a valid IRCC "
         "registered in ABSCH. For non-IRCC countries (e.g. Brazil), platform "
         "surfaces the country-specific equivalent and flags the distinction.",
         "EU User / Admin",
         "Authenticated",
         "ABSCH IRCC API endpoint",
         "ABSCH (IRCC search)",
         "Low (API call) / Medium (exception handling)",
         "IRCC data is legally sensitive. Must not be stored or transmitted "
         "in ways that compromise CBI.",
         "Y — IRCC is the foundational legal proof of compliant access",
         "High",
         "[SRS-REQ] IRCC verification must distinguish between IRCC-issuing "
         "countries and equivalent-certificate countries. For Brazil, "
         "platform must request SisGen registration receipt as the "
         "compliance proof."],

        ["Verification & Due Diligence",
         "Patent & Prior Art Screening (Biopiracy Check)",
         "Before a resource match is made, platform checks Lens.org and WIPO "
         "Patentscope for existing patents referencing the resource. Flags "
         "cases where EU IPR has been filed without documented ABS compliance "
         "— protecting Southern providers and informing EU user due diligence.",
         "EU User / Southern Provider / Admin",
         "Authenticated",
         "Lens.org (primary), WIPO Patentscope, Espacenet",
         "Espacenet, Lens.org",
         "Medium",
         "Flagging 'potential biopiracy' must be framed as advisory only. "
         "Platform carries no legal liability for screening results.",
         "Y — identifies existing IPR as part of due diligence",
         "High",
         "[SRS-REQ] Patent screening must be a mandatory step in the pre-match "
         "due diligence workflow, with results framed as advisory flags "
         "requiring legal review — not automatic deal blockers."],

        ["Verification & Due Diligence",
         "DSI Provenance Tracking",
         "For any digital sequence information (DSI) derived from a resource "
         "accessed through the platform, tracks the chain from physical sample "
         "(GGBN permit vocabulary) to sequence deposit (GenBank / ENA accession "
         "number). Creates an auditable link between the ABS agreement and "
         "downstream utilisation.",
         "EU User / Admin",
         "Authenticated",
         "GGBN Data Portal, GenBank (NCBI), ENA (EMBL-EBI), DDBJ",
         "GGBN Data Portal (provenance model)",
         "High",
         "DSI provenance data is legally sensitive under evolving CBD "
         "multilateral mechanism discussions. Must be treated with equivalent "
         "confidentiality to MTA data.",
         "Y — DSI provenance is a compliance frontier",
         "High",
         "[SRS-REQ] Platform must track GenBank/ENA accession numbers for "
         "any DSI derived from platform-mediated resource access, creating "
         "an auditable provenance chain from physical sample to sequence "
         "database record."],

        ["Verification & Due Diligence", "MTA Generation & Workflow",
         "Platform provides templates for Material Transfer Agreements (MTAs) "
         "based on GIZ model clauses (default), OpenMTA (open/research), and "
         "UBMTA (academic non-commercial). Custom GENE-LINK Partnership "
         "Agreement Template for commercial deals. E-signature integration "
         "for execution. Executed MTAs stored in encrypted vault.",
         "EU User / Southern Provider",
         "Authenticated (Verified tier minimum)",
         "GIZ ABS Contractual Clauses, OpenMTA template library, UBMTA, "
         "Custom GENE-LINK Agreement Template",
         "DocSend (access tracking), Intralinks (audit trail), "
         "Upwork (milestone contracts), MTAShare",
         "High",
         "Executed MTAs contain confidential commercial terms. "
         "AES-256 encryption at rest; TLS 1.3 in transit. Access-controlled "
         "per document by uploading party.",
         "Y — MTA execution is a core compliance and legal protection step",
         "High",
         "[SRS-REQ] Platform must provide GIZ model clauses as the default "
         "MTA starting point, with OpenMTA as the open-access option. "
         "Executed MTAs must be stored in an encrypted, access-controlled "
         "document vault."],

        ["Verification & Due Diligence",
         "Identity & Institution Verification",
         "Tiered verification of user identity and institutional affiliation "
         "as a precondition for accessing compliance workflows and initiating "
         "deal flows. Uses ORCID for researcher identity, OpenCorporates "
         "for EU company registration, and manual review for institutions "
         "in countries without open registries.",
         "All",
         "Platform-level prerequisite",
         "ORCID, OpenCorporates",
         "Airbnb (identity verification model), LinkedIn (credibility signals), "
         "Upwork (profile + history)",
         "High",
         "GDPR compliance required. Minimum data collection. OWASP-compliant "
         "authentication. MFA for Verified tier.",
         "Y — user identity verification is part of due diligence",
         "High",
         "[SRS-REQ] Platform must implement a four-tier access control model: "
         "Public / Authenticated / Verified / Admin. Verified tier requires "
         "institutional affiliation confirmation and is a prerequisite for "
         "initiating any ABS compliance or deal workflow."],
    ],

    "SECURITY": [
        ["Security", "Encrypted Document Vault",
         "Secure, encrypted storage for sensitive documents: executed MTAs, "
         "benefit-sharing agreements, IRCC records, SisGen receipts, and "
         "GGBN provenance records. Version-controlled and audit-trailed. "
         "EU data residency for EU user documents.",
         "EU User / Southern Provider",
         "Authenticated (Verified tier minimum)",
         "GGBN Data Standard (provenance metadata model)",
         "Intralinks (data room model), DocSend (access tracking), "
         "Egnyte for Life Sciences (GxP vault model)",
         "High",
         "AES-256 encryption at rest; TLS 1.3 in transit. Per-document "
         "access control lists. Immutable access audit logs. EU data "
         "residency for GDPR compliance.",
         "Y — document vault is the compliance record of the exchange",
         "High",
         "[SRS-REQ] Document vault must support: AES-256 encryption; "
         "TLS 1.3; per-document ACLs; immutable access audit logs; "
         "EU data residency for EU user documents."],

        ["Security", "Data Provenance & Integrity Tracking",
         "Every genetic resource record carries a provenance chain "
         "following the GGBN permit vocabulary (PIC, MAT, MTA fields). "
         "Cryptographic checksums on key compliance records to detect "
         "tampering. Immutable once created.",
         "All",
         "Authenticated",
         "GGBN Data Standard, ABSCH (IRCC reference), GenBank/ENA",
         "GGBN Data Portal (provenance model), GitHub (version control)",
         "High",
         "Provenance records must be immutable once created. Hash-based "
         "integrity checking on IRCC and compliance documents.",
         "Y — data provenance is the core of the compliance architecture",
         "High",
         "[SRS-REQ] All genetic resource records must carry GGBN permit "
         "vocabulary metadata. Platform must generate and store a "
         "cryptographic hash of key compliance documents at creation "
         "to support future tamper detection."],

        ["Security", "Secure Communication Layer",
         "Encrypted, channel-based communication for active partnerships. "
         "Integration with MS Teams (EU institutional default) and/or "
         "Rocket.Chat (open-source, data-sovereign option) rather than "
         "building a custom messaging system.",
         "EU User / Southern Provider",
         "Authenticated",
         "N/A (infrastructure-level)",
         "Microsoft Teams (primary), Rocket.Chat (alternative)",
         "Medium",
         "All messages containing agreement terms or compliance data must "
         "be end-to-end encrypted. Data residency must be configurable "
         "for partners with strict hosting requirements.",
         "N",
         "High",
         "[SRS-REQ] Platform secure communication module must integrate with "
         "MS Teams and offer Rocket.Chat as a data-sovereign alternative "
         "rather than building a custom messaging layer from scratch."],
    ],

    "AI / ML FEATURES": [
        ["AI / ML", "AI-Powered Matchmaking Engine",
         "ML recommendation system matching Southern providers with EU "
         "users based on: resource attributes (species, compound class, "
         "ecosystem), sector needs, ABS compliance readiness, geographic "
         "preferences, and historical partnership outcomes. Implements "
         "mutual-interest logic before surfacing a match.",
         "All",
         "Authenticated",
         "Custom Partner Registry, Custom Opportunity Pipeline, GBIF, Genesys",
         "Grip (AI recommendation engine), b2match (structured fields), "
         "Tinder/Hinge (mutual interest + profile depth)",
         "High",
         "Model training data must not expose individual user behaviour "
         "without consent. Explainability requirement: users should "
         "understand why a match was recommended.",
         "N (but supports due diligence by surfacing compliance status)",
         "High",
         "[SRS-REQ] Matchmaking engine must implement mutual-interest logic "
         "(both parties signal interest before full match is revealed), "
         "ABS-regime-aware filtering, and explainable recommendation "
         "summaries."],

        ["AI / ML", "NLP Resource Description & Tagging",
         "Natural language processing to auto-tag resource descriptions, "
         "institutional profiles, and opportunity listings with standardised "
         "taxonomic codes (GBIF backbone), sector classifications, and "
         "ABS-relevance indicators. Reduces manual tagging burden and "
         "improves search relevance.",
         "Admin / All (benefits all users)",
         "Platform-level (background processing)",
         "GBIF (taxonomic backbone), Catalogue of Life, GGBN permit vocab",
         "GitHub (tagging model inspiration)",
         "Medium",
         "NLP models must be trained on multilingual text given the "
         "global user base. Output tags must be human-reviewable before "
         "being surfaced to match algorithms.",
         "N",
         "High",
         "[SRS-REQ] Platform must implement NLP-based auto-tagging for all "
         "resource descriptions and institutional profiles using GBIF "
         "taxonomic backbone and GGBN permit vocabulary as the tagging "
         "schema."],

        ["AI / ML", "AI-Assisted ABS Compliance Guidance",
         "LLM-based conversational assistant that interprets user inputs "
         "(e.g. 'I want to access neem extract from India for a new "
         "cosmetic formulation') and surfaces the relevant national ABS "
         "compliance pathway, permit requirements, documentation checklist, "
         "and Nagoya Protocol Hub country guide. Backed by live ABSCH "
         "data and AbsMaster.",
         "EU User / Southern Provider",
         "Authenticated",
         "ABSCH API, CGHSS AbsMaster, SisGen guidance, NBA e-Filing, "
         "Nagoya Protocol Hub, GBIF",
         "Nagoya Protocol Hub (guided walkthrough model), CTCN (request-"
         "response model)",
         "High",
         "AI guidance must include a mandatory disclaimer that outputs "
         "are informational only and do not constitute legal advice. "
         "All guidance must cite its source data.",
         "Y — core compliance guidance function",
         "High",
         "[SRS-REQ] Platform must include an LLM-based ABS compliance "
         "assistant that interprets natural language queries and surfaces "
         "the correct national pathway, with source citations and a "
         "clear legal disclaimer."],

        ["AI / ML", "Conversational Onboarding Bot",
         "Guided conversational interface for new users (EU companies and "
         "Southern institutions) to complete profiles, understand platform "
         "obligations, and identify their first actions. Reduces drop-off "
         "during the onboarding process by making it interactive rather "
         "than form-based.",
         "All (new users)",
         "Public (initial) / Authenticated (profile completion)",
         "N/A (UX-level feature)",
         "Grip (guided onboarding with tagging), LinkedIn (progressive "
         "profile), Hinge (prompt-based profile completion)",
         "Medium",
         "Onboarding data collected by the bot must be subject to the "
         "same GDPR treatment as form-based data collection.",
         "N",
         "High",
         "[SRS-REQ] Platform onboarding must be conversational rather than "
         "purely form-based — an AI bot should guide new users through "
         "profile completion, role selection, and first-action identification."],

        ["AI / ML", "Predictive Partnership Fit Scoring",
         "ML model scoring provider-user compatibility based on: resource "
         "match quality, ABS compliance readiness of provider, sector "
         "alignment, geographic feasibility, and (over time) historical "
         "partnership outcome data. Score surfaced to admin and as an "
         "optional ranking signal.",
         "Admin / (optionally) EU User",
         "Authenticated (Admin-visible by default; optionally shared)",
         "Custom Partner Registry, Custom Opportunity Pipeline",
         "AngelList (investment fit scoring), Dealroom (ecosystem analytics)",
         "High",
         "Fit scores must not be shared with Southern providers in ways "
         "that create discriminatory ranking effects. Score methodology "
         "must be documented and auditable.",
         "N",
         "Medium",
         "[SRS-REQ] Partnership fit scoring model must be documented, "
         "auditable, and free from discriminatory proxies. Score methodology "
         "must be disclosed to admin users."],

        ["AI / ML", "Intelligent Regulatory Alert System",
         "Monitors ABSCH, national government portals, and legal news "
         "sources for regulatory changes (new ABS laws, IRCC policy updates, "
         "national permit fee changes) and automatically alerts platform "
         "users who have active partnerships or saved searches in the "
         "affected provider country.",
         "All",
         "Authenticated",
         "ABSCH API, CGHSS AbsMaster, national portal monitoring",
         "Power BI (alerts model), Confluence (update notifications)",
         "High",
         "Alert system must not expose non-public regulatory intelligence "
         "or provide competitive advantage to specific users.",
         "Y — regulatory change monitoring is part of ongoing due diligence",
         "Medium",
         "[SRS-REQ] Platform must implement automated regulatory alerts "
         "for changes to ABS requirements in provider countries where users "
         "have active engagements or saved interests."],

        ["AI / ML", "AI-Assisted MTA Clause Generation",
         "LLM-assisted drafting of MTA and benefit-sharing agreement clauses "
         "based on: resource type, provider country jurisdiction, user intent "
         "(research/commercial), and existing model clauses from GIZ, "
         "OpenMTA, and UBMTA libraries. Reduces legal cost and negotiation "
         "friction for first-time users.",
         "EU User / Southern Provider",
         "Authenticated (Verified tier)",
         "GIZ ABS Contractual Clauses, OpenMTA, UBMTA, Custom GENE-LINK "
         "Partnership Agreement Template",
         "DocSend (document workflow), Intralinks (agreement execution)",
         "High",
         "AI-generated clauses must be clearly labelled as draft suggestions "
         "requiring legal review before execution. Platform carries no "
         "liability for AI-generated legal text.",
         "Y — MTA generation is a core due diligence and compliance step",
         "High",
         "[SRS-REQ] AI MTA drafting tool must: cite the model clause source "
         "for each suggested clause, label all outputs as 'draft for legal "
         "review', and require explicit user confirmation before any clause "
         "is included in a final agreement."],

        ["AI / ML", "Ecosystem Analytics & Insights Dashboard",
         "AI-generated insights on platform activity: demand signals by "
         "resource type and sector, compliance bottlenecks by country, "
         "partnership pipeline health, geographic coverage gaps, and "
         "matching efficiency metrics. Available to admin and (aggregated, "
         "anonymised) to platform partners.",
         "Admin / (aggregated) All",
         "Admin-Only (detailed) / Authenticated (aggregated summary)",
         "Custom Partner Registry, Custom Opportunity Pipeline, ABSCH",
         "Power BI (dashboarding), Dealroom (ecosystem mapping), "
         "Crunchbase (discovery data)",
         "Medium",
         "Dashboard data must be aggregated and anonymised before sharing "
         "with non-admin users to prevent identification of individual "
         "user behaviour.",
         "N",
         "Medium",
         "[SRS-REQ] Platform admin dashboard must provide real-time "
         "analytics on: partnership pipeline stages, compliance completion "
         "rates, geographic coverage, and demand signal trends by sector "
         "and resource type."],
    ],
}


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------

def _write_data_source_sheet(ws, rows, header_fill, tab_subtitle, n_hdr_cols=18):
    """Write a standard II-x data source sheet."""
    _banner(ws, ws.title.split(". ", 1)[-1] if ". " in ws.title else ws.title,
            tab_subtitle)
    _header(ws, DS_COLS, row=3, fill_hex=header_fill)
    for i, r in enumerate(rows, start=4):
        _row(ws, r, i, shade=(i % 2 == 0))
        ws.row_dimensions[i].height = 80
    w = [26, 45, 45, 18, 12, 18, 38, 22, 14, 28, 16, 14,
         12, 22, 18, 10, 45, 50]
    _widths(ws, w)
    ws.freeze_panes = "B4"


def build_workbook(output_path: str) -> None:
    """Build and save the merged GENE-LINK Platform Scan workbook."""
    wb = Workbook()

    # ── 0. Instructions ──────────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "0. Instructions"
    _banner(ws0,
            "GENE-LINK Platform Scan — Platform Survey & Mapping (Merged v0.2)",
            "Merges GENE-LINK Platform Team v0.1 with Dominique's Platform "
            "Ecosystem Scan Tool. All source information preserved.")
    ws0.cell(row=3, column=1, value="SECTION").font = _font(
        bold=True, color=WHITE, size=9)
    ws0.cell(row=3, column=1).fill = _fill(NAVY)
    ws0.cell(row=3, column=1).alignment = _center()
    ws0.cell(row=3, column=1).border = _border()
    ws0.cell(row=3, column=2, value="DETAIL").font = _font(
        bold=True, color=WHITE, size=9)
    ws0.cell(row=3, column=2).fill = _fill(NAVY)
    ws0.cell(row=3, column=2).alignment = _center()
    ws0.cell(row=3, column=2).border = _border()
    for ri, (sec, det) in enumerate(INSTRUCTIONS, start=4):
        bg = BLUE_PALE if ri % 2 == 0 else WHITE
        c1 = ws0.cell(row=ri, column=1, value=sec)
        c1.font = _font(bold=True, color=NAVY, size=9)
        c1.fill, c1.alignment, c1.border = _fill(bg), _left(), _border()
        c2 = ws0.cell(row=ri, column=2, value=det)
        c2.font = _font(size=9)
        c2.fill, c2.alignment, c2.border = _fill(bg), _left(), _border()
        ws0.row_dimensions[ri].height = 54
    ws0.column_dimensions["A"].width = 34
    ws0.column_dimensions["B"].width = 115

    # ── I-A. Existing Platforms ──────────────────────────────────────────────
    ws1a = wb.create_sheet("I-A. Existing Platforms")
    _banner(ws1a,
            "PART I-A — EXISTING PLATFORMS (38 platforms)",
            "Covers matchmaking, collaboration, document infrastructure, "
            "marketplace/trust, analytics, compliance, capacity development, "
            "e-learning, innovation networks, and UX reference platforms.")
    _header(ws1a, PLAT_COLS, row=3, fill_hex=BLUE_MID)
    ws1a.row_dimensions[3].height = 42

    row_n = 4
    section_fills = {
        "MATCHMAKING": INDIGO_PALE,
        "COMMUNITY & COLLABORATION": TEAL_PALE,
        "DOCUMENT & DEAL INFRASTRUCTURE": AMBER_PALE,
        "MARKETPLACE & TRUST MODELS": BROWN_PALE,
        "PROFESSIONAL & ACADEMIC NETWORKS": PALE_GREEN,
        "ANALYTICS, DATA & KNOWLEDGE": PURPLE_PALE,
        "REGULATORY COMPLIANCE": "FCE4EC",
        "INNOVATION NETWORKS": BLUE_PALE,
        "CAPACITY DEVELOPMENT": PALE_GREEN,
        "E-LEARNING": "E1F5FE",
        "UX & BEHAVIOURAL REFERENCE": "FFF8E1",
    }
    section_header_fills = {
        "MATCHMAKING": INDIGO,
        "COMMUNITY & COLLABORATION": TEAL,
        "DOCUMENT & DEAL INFRASTRUCTURE": AMBER,
        "MARKETPLACE & TRUST MODELS": BROWN,
        "PROFESSIONAL & ACADEMIC NETWORKS": MID_GREEN,
        "ANALYTICS, DATA & KNOWLEDGE": PURPLE,
        "REGULATORY COMPLIANCE": "880E4F",
        "INNOVATION NETWORKS": BLUE_MID,
        "CAPACITY DEVELOPMENT": DARK_GREEN,
        "E-LEARNING": "0277BD",
        "UX & BEHAVIOURAL REFERENCE": "F57F17",
    }
    for section_name, section_rows in PLAT_ROWS.items():
        _section(ws1a, f"▸  {section_name}",
                 row_n, fill_hex=section_fills.get(section_name, BLUE_PALE),
                 ncols=21)
        row_n += 1
        for r in section_rows:
            _row(ws1a, r, row_n, shade=(row_n % 2 == 0))
            ws1a.row_dimensions[row_n].height = 80
            row_n += 1

    plat_widths = [26, 32, 20, 9, 42, 16, 22, 35, 28, 22,
                   38, 35, 35, 35, 38, 32, 14, 12, 12, 14, 52]
    _widths(ws1a, plat_widths)
    ws1a.freeze_panes = "B4"

    # ── I-B. Architecture Notes ──────────────────────────────────────────────
    ws1b = wb.create_sheet("I-B. Architecture Notes")
    _banner(ws1b,
            "PART I-B — PLATFORM ARCHITECTURE NOTES",
            "UX/technical deep-dive on each platform: onboarding flows, "
            "matching algorithms, data models, design features to adopt, "
            "and design failures to avoid. Cross-reference by Platform Name "
            "to I-A. Source: Dominique (Platform Ecosystem Scan Tool).")
    _header(ws1b, ARCH_COLS, row=3, fill_hex=INDIGO)
    for i, r in enumerate(ARCH_ROWS, start=4):
        _row(ws1b, r, i, shade=(i % 2 == 0))
        ws1b.row_dimensions[i].height = 70
    arch_widths = [22, 28, 40, 32, 38, 35, 32,
                   32, 28, 14, 14, 28, 38, 38, 14, 14, 28]
    _widths(ws1b, arch_widths)
    ws1b.freeze_panes = "B4"

    # ── II-A. Regulatory Sources ─────────────────────────────────────────────
    ws2a = wb.create_sheet("II-A. Regulatory Sources")
    _write_data_source_sheet(
        ws2a, REG_ROWS, MID_GREEN,
        "Regulatory ABS data sources. Goal: identify data sources for the "
        "compliance and due diligence backend. Includes ABSCH, national "
        "portals, patent databases, and global ABS datasets.")

    # ── II-B. Company Sources ────────────────────────────────────────────────
    ws2b = wb.create_sheet("II-B. Company Sources")
    _write_data_source_sheet(
        ws2b, COMP_ROWS, AMBER,
        "EU company and innovation databases. Goal: inform EU user "
        "matchmaking, company profiling, and due diligence backend.")

    # ── II-C. Resource & Biodiversity ───────────────────────────────────────
    ws2c = wb.create_sheet("II-C. Resource & Biodiversity")
    _write_data_source_sheet(
        ws2c, BIO_ROWS, PURPLE,
        "Resource, biodiversity, and genomic databases — including DSI "
        "databases (GenBank, ENA, DDBJ), species occurrence (GBIF), "
        "plant genetic resources (Genesys), and protected areas (WDPA).")

    # ── II-D. Southern Institution DBs ──────────────────────────────────────
    ws2d = wb.create_sheet("II-D. Southern Institution DBs")
    _write_data_source_sheet(
        ws2d, SOUTH_ROWS, TEAL,
        "Southern institution databases and directories. Goal: populate "
        "verified provider institution profiles and support matchmaking.")

    # ── II-E. Content & Knowledge Assets ────────────────────────────────────
    ws2e = wb.create_sheet("II-E. Content & Knowledge")
    _banner(ws2e,
            "PART II-E — CONTENT & KNOWLEDGE ASSETS",
            "Merged E-Learning resources and Knowledge/Legal assets. "
            "Section 1: E-Learning courses and training modules. "
            "Section 2: Legal templates, frameworks, and standards. "
            "Dedup notes flag where a source's data-API dimension is "
            "captured in II-A through II-D.")
    _header(ws2e, CONTENT_COLS, row=3, fill_hex=DARK_GREEN)

    row_n = 4
    content_section_fills = {
        "E-LEARNING COURSES & TRAINING MODULES": PALE_GREEN,
        "KNOWLEDGE & LEGAL ASSETS": BLUE_PALE,
    }
    content_section_headers = {
        "E-LEARNING COURSES & TRAINING MODULES": MID_GREEN,
        "KNOWLEDGE & LEGAL ASSETS": BLUE_MID,
    }
    for section_name, section_rows in CONTENT_ROWS.items():
        _section(ws2e, f"▸  {section_name}",
                 row_n,
                 fill_hex=content_section_fills.get(section_name, PALE_GREEN),
                 ncols=24)
        row_n += 1
        for r in section_rows:
            _row(ws2e, r, row_n, shade=(row_n % 2 == 0))
            ws2e.row_dimensions[row_n].height = 75
            row_n += 1

    content_widths = [30, 16, 32, 24, 20, 42, 9, 9, 9, 9, 9,
                      22, 12, 10, 14, 16, 12, 14, 10, 18, 20, 10, 38, 50]
    _widths(ws2e, content_widths)
    ws2e.freeze_panes = "B4"

    # ── II-F. In-Country Support ─────────────────────────────────────────────
    ws2f = wb.create_sheet("II-F. In-Country Support")
    _banner(ws2f,
            "PART II-F — IN-COUNTRY SUPPORT ECOSYSTEM",
            "One row per organisation, per country. Covers accelerators, "
            "legal/IP services, commercialisation support, and ABS advisory "
            "services in provider countries. Template to be populated. "
            "Source: Dominique (Platform Ecosystem Scan Tool).")
    _header(ws2f, INCOUNTRY_COLS, row=3, fill_hex=BROWN)
    # Write 5 empty template rows
    for i in range(4, 9):
        _row(ws2f, [""] * len(INCOUNTRY_COLS), i, shade=(i % 2 == 0))
        ws2f.row_dimensions[i].height = 40
    incountry_widths = [14, 28, 30, 20, 18, 38, 10, 10, 10, 10, 18,
                        16, 14, 22, 28, 30, 22, 12, 14, 14, 30]
    _widths(ws2f, incountry_widths)
    ws2f.freeze_panes = "B4"

    # ── III. Functionalities & Requirements ──────────────────────────────────
    ws3 = wb.create_sheet("III. Functionalities")
    _banner(ws3,
            "PART III — FUNCTIONALITIES & PLATFORM REQUIREMENTS",
            "Four sections: Access & Discovery | Verification & Due Diligence "
            "| Security | AI / ML Features. Each row maps to a [SRS-REQ] "
            "statement for the System Requirements Specification.")

    func_section_fills = {
        "ACCESS & DISCOVERY":           BLUE_PALE,
        "VERIFICATION & DUE DILIGENCE": AMBER_PALE,
        "SECURITY":                     PURPLE_PALE,
        "AI / ML FEATURES":             PALE_GREEN,
    }
    func_section_headers = {
        "ACCESS & DISCOVERY":           BLUE_MID,
        "VERIFICATION & DUE DILIGENCE": AMBER,
        "SECURITY":                     PURPLE,
        "AI / ML FEATURES":             MID_GREEN,
    }

    row_n = 3
    for section_name, section_rows in FUNC_ROWS.items():
        _section(ws3, f"▸  {section_name}",
                 row_n,
                 fill_hex=func_section_fills.get(section_name, BLUE_PALE),
                 ncols=12)
        row_n += 1
        _header(ws3, FUNC_COLS, row=row_n,
                fill_hex=func_section_headers.get(section_name, BLUE_MID))
        row_n += 1
        for r in section_rows:
            _row(ws3, r, row_n, shade=(row_n % 2 == 0))
            ws3.row_dimensions[row_n].height = 90
            row_n += 1

    func_widths = [20, 28, 52, 18, 16, 45, 35, 12, 42, 18, 10, 58]
    _widths(ws3, func_widths)
    ws3.freeze_panes = "B5"

    # ── Tab colours ──────────────────────────────────────────────────────────
    tab_colors = {
        "0. Instructions":              "1A237E",
        "I-A. Existing Platforms":      "1565C0",
        "I-B. Architecture Notes":      "283593",
        "II-A. Regulatory Sources":     "2E7D32",
        "II-B. Company Sources":        "E65100",
        "II-C. Resource & Biodiversity": "4A148C",
        "II-D. Southern Institution DBs": "004D40",
        "II-E. Content & Knowledge":    "1B5E20",
        "II-F. In-Country Support":     "4E342E",
        "III. Functionalities":         "311B92",
    }
    for title, color in tab_colors.items():
        if title in wb.sheetnames:
            wb[title].sheet_properties.tabColor = color

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    OUTPUT = "./GENE_LINK_Platform_Scan_v02.xlsx"
    build_workbook(OUTPUT)
